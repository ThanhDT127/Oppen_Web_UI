"""
Script tạo file Excel checklist tính năng Open WebUI
Ghi vào sheet 'list tính năng' trong file List function.xlsx
"""
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import os

EXCEL_PATH = r"C:\Code\openwebui_fetch\Oppen_Web_UI\docs\List function.xlsx"

# ============================================================
# DATA
# ============================================================
HEADER_ROW = ["STT", "Module", "Nhóm tính năng", "Tính năng cụ thể",
              "Hướng dẫn sử dụng / Mô tả tính năng", "Câu lệnh ví dụ",
              "Trạng thái tính năng", "Kết quả test", "Ghi chú"]

TITLE = "CHECK LIST TÍNH NĂNG OPEN WEBUI"

# Data: list of (module_header, rows)
# Each row: (stt, nhom_tinh_nang, tinh_nang, huong_dan, vi_du, trang_thai, ket_qua, ghi_chu)
MODULES = [
    ("I. PHÂN QUYỀN & QUẢN LÝ NGƯỜI DÙNG", [
        ("1", "Đăng ký & Đăng nhập", "Đăng ký tài khoản bằng email/password",
         "Người dùng truy cập http://<server>:3000, chọn 'Sign Up', điền tên, email, mật khẩu", "",
         "Đã có", "OK", "ENABLE_SIGNUP=true"),
        ("2", "", "Đăng nhập bằng email/password",
         "Nhập email + password → nhận JWT token", "",
         "Đã có", "OK", ""),
        ("3", "", "Hỗ trợ OAuth/SSO",
         "Cho phép đăng nhập qua AD/LDAP/Google/Microsoft", "",
         "Hạ tầng có", "Chưa test", "Cần cấu hình OAUTH_* env"),
        ("4", "Phân quyền", "3 cấp bậc: Admin, User, Pending",
         "Admin: quản lý toàn bộ (users, models, knowledge, settings)\nUser: sử dụng chat, upload file, tạo knowledge cá nhân\nPending: chờ admin duyệt",
         "", "Đã có", "OK", ""),
        ("5", "", "Quản lý user (Admin)",
         "Admin có thể: thêm/xoá user, đổi role, reset password, xem danh sách user",
         "Admin → Settings → Users", "Đã có", "OK", ""),
        ("6", "", "Access Control trên Knowledge",
         "Mỗi Knowledge Collection có thể giới hạn quyền truy cập theo user/group",
         "Cấu hình trong Knowledge Settings", "Đã có", "OK", "Hỗ trợ JSON access_control"),
        ("7", "", "Access Control trên Model",
         "Admin có thể giới hạn model nào user nào được dùng",
         "Admin → Settings → Models", "Đã có", "OK", ""),
        ("8", "Quản lý log (Admin)", "Log hoạt động API requests",
         "Middleware ghi log mọi request: user, model, tokens, cost, timestamp",
         "Xem file logs/middleware.requests.log", "Đã có", "OK", ""),
        ("9", "", "Dashboard quản trị",
         "Xem tổng quan chi phí, số request, top users, top models",
         "http://<server>:5000/dashboard", "Đã có", "OK", "Middleware dashboard"),
    ]),
    ("II. CHAT AI ĐA MÔ HÌNH", [
        ("1", "OpenAI GPT-5 Series", "Chat GPT-5 (Flagship)",
         "Model mạnh nhất, đa phương thức (text + hình ảnh + audio)",
         "Phân tích hình ảnh này và cho nhận xét", "Đã có", "OK", "chat-gpt-5"),
        ("2", "", "Chat GPT-5 Mini",
         "Nhanh hơn GPT-5, phù hợp tác vụ thông thường",
         "Tóm tắt email này", "Đã có", "OK", "chat-gpt-5-mini"),
        ("3", "", "Chat GPT-5 Nano",
         "Nhẹ nhất, chi phí thấp nhất",
         "Dịch câu này sang tiếng Anh", "Đã có", "OK", "chat-gpt-5-nano"),
        ("4", "OpenAI GPT-4o Series", "Chat GPT-4o",
         "Multimodal nhanh, hỗ trợ hình ảnh, âm thanh",
         "Mô tả nội dung hình ảnh", "Đã có", "OK", "chat-gpt-4o"),
        ("5", "", "Chat GPT-4o Mini",
         "Nhanh, rẻ, đủ tốt cho hầu hết tác vụ",
         "Viết email trả lời khách hàng", "Đã có", "OK", "chat-gpt-4o-mini"),
        ("6", "OpenAI GPT-4.1 Series", "Chat GPT-4.1 (1M context)",
         "Context window 1 triệu token, xử lý tài liệu siêu dài",
         "Phân tích toàn bộ báo cáo 200 trang này", "Đã có", "OK", "chat-gpt-4.1"),
        ("7", "", "Chat GPT-4.1 Mini",
         "1M context, chi phí thấp hơn",
         "Tóm tắt tài liệu dài", "Đã có", "OK", "chat-gpt-4.1-mini"),
        ("8", "", "Chat GPT-4.1 Nano",
         "1M context, chi phí thấp nhất",
         "Trích xuất thông tin từ file lớn", "Đã có", "OK", "chat-gpt-4.1-nano"),
        ("9", "Google Gemini 2.5 Series", "Chat Gemini 2.5 Pro",
         "Reasoning mạnh, hỗ trợ tiếng Việt tốt",
         "Giải thích khái niệm quantum computing", "Đã có", "OK", "chat-gemini-2.5-pro"),
        ("10", "", "Chat Gemini 2.5 Flash",
         "Nhanh, cân bằng giữa tốc độ và chất lượng",
         "Viết code Python sort danh sách", "Đã có", "OK", "chat-gemini-2.5-flash"),
        ("11", "", "Chat Gemini 2.5 Flash Lite",
         "Nhẹ nhất Gemini 2.5, chi phí rẻ",
         "Kiểm tra ngữ pháp câu này", "Đã có", "OK", "chat-gemini-2.5-flash-lite"),
        ("12", "Google Gemini 2.0 Series", "Chat Gemini 2.0 Flash",
         "Thế hệ trước, vẫn ổn định",
         "Chat thông thường", "Đã có", "OK", "chat-gemini-2.0-flash"),
        ("13", "", "Chat Gemini 2.0 Flash Lite",
         "Nhẹ, nhanh, rẻ",
         "Trả lời câu hỏi đơn giản", "Đã có", "OK", "chat-gemini-2.0-flash-lite"),
        ("14", "Google Gemini 3 Series", "Chat Gemini 3 Pro (Flagship)",
         "Model Flagship mới nhất của Google",
         "Phân tích dữ liệu phức tạp", "Đã có", "OK", "chat-gemini-3-pro"),
        ("15", "Tính năng chat chung", "Streaming response realtime",
         "Response hiển thị từng token realtime, không phải chờ toàn bộ",
         "", "Đã có", "OK", ""),
        ("16", "", "Lưu lịch sử hội thoại",
         "Toàn bộ chat được lưu trong PostgreSQL, có thể xem lại",
         "", "Đã có", "OK", "Table: chat"),
        ("17", "", "Pin / Archive hội thoại",
         "Ghim hội thoại quan trọng hoặc lưu trữ",
         "", "Đã có", "OK", ""),
        ("18", "", "Chia sẻ hội thoại qua link",
         "Tạo link công khai để người khác xem hội thoại",
         "Share icon → Copy link", "Đã có", "OK", ""),
        ("19", "", "Folder tổ chức hội thoại",
         "Sắp xếp hội thoại vào folder theo chủ đề",
         "", "Đã có", "OK", ""),
        ("20", "", "Gắn tag phân loại",
         "Tag hội thoại để dễ tìm kiếm",
         "", "Đã có", "OK", ""),
        ("21", "", "Chuyển model giữa chừng",
         "Đổi model trong cùng 1 hội thoại, không mất context",
         "Chọn model khác từ dropdown", "Đã có", "OK", ""),
        ("22", "", "Multimodal input (ảnh + text)",
         "Gửi hình ảnh kèm text trong chat để AI phân tích",
         "Kéo thả hoặc paste hình vào chat", "Đã có", "OK", "GPT-4o, Gemini"),
    ]),
    ("III. TẠO ẢNH AI (IMAGE GENERATION)", [
        ("1", "OpenAI DALL-E 3", "Tạo ảnh từ mô tả text",
         "Chọn model img-gpt-dalle-3, nhập mô tả → nhận ảnh 1024x1024",
         "Vẽ hình con mèo đang đọc sách", "Đã có", "OK", "img-gpt-dalle-3"),
        ("2", "Google Gemini Image", "Tạo ảnh nhanh (Gemini Flash)",
         "Chọn img-gemini-flash, nhập mô tả → nhận ảnh nhanh",
         "Tạo logo công ty màu xanh", "Đã có", "OK", "img-gemini-flash, 1024px"),
        ("3", "", "Tạo ảnh chất lượng cao (Gemini Pro)",
         "Chọn img-gemini-pro, nhập mô tả → ảnh chất lượng cao lên đến 4K",
         "Thiết kế poster quảng cáo đèn LED", "Đã có", "OK", "img-gemini-pro, 4K"),
    ]),
    ("IV. GIỌNG NÓI (VOICE)", [
        ("1", "Text-to-Speech (TTS)", "Chuyển text thành giọng nói",
         "AI đọc nội dung response bằng giọng nói tự nhiên",
         "Nhấn icon 🔊 trên response", "Đã có", "OK", "tts-gpt-4o-mini"),
        ("2", "Speech-to-Text (STT)", "Nhập liệu bằng giọng nói",
         "Nói vào mic → AI chuyển thành text → gửi prompt",
         "Nhấn icon 🎤 trong chat", "Đã có", "OK", "stt-gpt-4o"),
        ("3", "", "Phiên bản nhẹ STT",
         "Nhận dạng giọng nói nhanh, chi phí thấp",
         "", "Đã có", "OK", "stt-gpt-4o-mini"),
    ]),
    ("V. RAG – CƠ SỞ TRI THỨC (KNOWLEDGE BASE)", [
        ("1", "Quản lý Knowledge", "Tạo Knowledge Collection",
         "Workspace → Knowledge → Create. Đặt tên, mô tả cho bộ tri thức",
         "", "Đã có", "OK", ""),
        ("2", "", "Upload file vào Knowledge",
         "Upload PDF, DOCX, TXT, CSV, MD, HTML, Excel. Tối đa 50MB/file, 10 file/lần",
         "Kéo thả file vào Knowledge", "Đã có", "OK", "RAG_FILE_MAX_SIZE=50"),
        ("3", "", "Xoá file khỏi Knowledge",
         "Xoá file → tự động xoá vector embeddings tương ứng",
         "", "Đã có", "OK", "CASCADE delete"),
        ("4", "", "Phân quyền truy cập Knowledge",
         "Giới hạn ai được truy cập Knowledge nào",
         "Knowledge Settings → Access Control", "Đã có", "OK", ""),
        ("5", "Sử dụng trong chat", "Gọi Knowledge bằng # (hashtag)",
         "Trong chat, gõ # rồi chọn Knowledge Collection",
         "#tai-lieu-noi-bo Chính sách nghỉ phép?", "Đã có", "OK", ""),
        ("6", "", "Attach file trực tiếp vào chat",
         "Kéo thả file vào chat, AI đọc và trả lời ngay (không lưu vào Knowledge)",
         "Kéo file PDF vào chat box", "Đã có", "OK", "Xử lý tạm thời"),
        ("7", "", "Gán Knowledge mặc định vào Model",
         "Admin gán Knowledge cho model → user dùng model đó tự động có context",
         "Admin → Models → Knowledge", "Đã có", "OK", ""),
        ("8", "Tìm kiếm thông minh", "Hybrid Search (BM25 + Vector)",
         "Kết hợp keyword matching (chính xác) + semantic search (ngữ nghĩa)",
         "", "Đã có", "OK", "ENABLE_RAG_HYBRID_SEARCH=true"),
        ("9", "", "HNSW Vector Index",
         "Tìm kiếm vector nhanh (approximate nearest neighbor), scale được triệu documents",
         "", "Đã có", "OK", "m=16, ef_construction=64"),
        ("10", "", "Citation – trích dẫn nguồn",
         "AI trích dẫn tên file và trang nguồn trong câu trả lời",
         "", "Đã có", "OK", ""),
        ("11", "Embedding", "Multilingual embedding (chạy local)",
         "Model chạy local, hỗ trợ 50+ ngôn ngữ (bao gồm tiếng Việt)",
         "", "Đã có", "OK", "paraphrase-multilingual-MiniLM-L12-v2"),
        ("12", "", "Bảo mật: Embedding không gửi ra ngoài",
         "Dữ liệu nội bộ KHÔNG rời khỏi server khi embedding",
         "", "Đã có", "OK", "Bảo mật dữ liệu nội bộ"),
        ("13", "Định dạng file hỗ trợ", "PDF",
         "Extract text từ PDF, hỗ trợ OCR", "", "Đã có", "OK", ""),
        ("14", "", "Word (.docx)", "Extract text từ file Word", "", "Đã có", "OK", ""),
        ("15", "", "Excel (.xlsx)", "Extract text từ file Excel", "", "Đã có", "OK", ""),
        ("16", "", "Text (.txt, .csv, .md)", "Đọc trực tiếp", "", "Đã có", "OK", ""),
        ("17", "", "HTML", "Extract text từ trang web", "", "Đã có", "OK", ""),
        ("18", "", "URL web", "Paste URL → hệ thống fetch và index nội dung", "", "Đã có", "OK", ""),
        ("19", "", "YouTube", "Paste URL YouTube → extract transcript", "", "Đã có", "OK", ""),
    ]),
    ("VI. CÔNG CỤ MỞ RỘNG (CUSTOM TOOLS)", [
        ("1", "Xuất Excel", "Trích xuất dữ liệu hội thoại → file Excel (.xlsx)",
         "Nhấn icon Action → 'Xuất Excel'. Hệ thống chuẩn hoá nội dung → bảng Markdown → file Excel có format, filter, freeze header",
         "Chat bảng biểu → nhấn icon xuất", "Đã có", "OK", "tool excel.py"),
        ("2", "", "Wizard UI với progress bar",
         "Modal hiển thị 3 bước: chuẩn hoá → tạo file → tải xuống",
         "", "Đã có", "OK", "UX thân thiện"),
        ("3", "", "Tự nhận dạng kiểu dữ liệu",
         "Nhận dạng: số, ngày tháng, phần trăm, tiền tệ VNĐ/$/ €",
         "", "Đã có", "OK", ""),
        ("4", "Xuất PDF", "Xuất nội dung hội thoại → file PDF",
         "Nhấn icon Action → 'Xuất PDF'",
         "", "Đã có", "OK", "tool pdf.py"),
        ("5", "Xuất DOCX", "Xuất nội dung hội thoại → file Word (.docx)",
         "Nhấn icon Action → 'Xuất DOCX'",
         "", "Đã có", "OK", "tool docx.py"),
        ("6", "Custom Functions", "Framework tạo function Python tuỳ chỉnh",
         "Admin upload code Python → chạy trong Open WebUI",
         "", "Đã có", "Chưa test", "Hạ tầng sẵn sàng"),
        ("7", "Custom Tools", "Framework tạo tool cho AI (function calling)",
         "Admin tạo tools → AI tự gọi khi cần",
         "", "Đã có", "Chưa test", "Hạ tầng sẵn sàng"),
    ]),
    ("VII. KIỂM SOÁT CHI PHÍ & QUOTA", [
        ("1", "Quota người dùng", "Giới hạn chi phí / user / tháng",
         "Mỗi user có limit_usd trong users.json. Hết quota → từ chối request (HTTP 429)",
         "", "Đã có", "OK", ""),
        ("2", "", "Cảnh báo khi gần hết quota",
         "User nhận thông báo cảnh báo khi sử dụng gần hết hạn mức",
         "", "Đã có", "OK", ""),
        ("3", "Sub-keys API", "Cấp API key riêng cho từng user/nhóm",
         "Mỗi user/nhóm có sub-key riêng biệt, quản lý trong users.json",
         "", "Đã có", "OK", ""),
        ("4", "Cost tracking", "Ghi log chi phí từng request",
         "Log: model, input_tokens, output_tokens, cost_usd, user, timestamp",
         "", "Đã có", "OK", "middleware.requests.log"),
        ("5", "", "Bảng giá riêng cho từng model",
         "Cập nhật bảng giá trong prices.json, tự tính cost",
         "", "Đã có", "OK", "prices.json"),
        ("6", "Dashboard", "Báo cáo chi phí trực quan",
         "Dashboard web: tổng chi phí, theo user, theo model, theo ngày/tuần/tháng",
         "http://<server>:5000/dashboard", "Đã có", "OK", ""),
    ]),
    ("VIII. DATABASE & HẠ TẦNG", [
        ("1", "PostgreSQL", "Database chính (PostgreSQL 16)",
         "Lưu toàn bộ data: users, chats, knowledge, vectors. 32 tables, 65 indexes",
         "", "Đã có", "OK", ""),
        ("2", "", "PGVector extension (v0.8.0)",
         "Vector similarity search cho RAG, HNSW index",
         "", "Đã có", "OK", ""),
        ("3", "", "Persistent storage (Docker volume)",
         "Data được giữ khi restart container",
         "", "Đã có", "OK", "Volume: postgres_data"),
        ("4", "Docker", "4 containers orchestrated",
         "PostgreSQL + LiteLLM + Middleware + Open WebUI. Docker Compose quản lý",
         "docker compose up -d", "Đã có", "OK", ""),
        ("5", "", "Health checks & Auto-restart",
         "Tự kiểm tra sức khoẻ và restart nếu container lỗi hoặc server reboot",
         "", "Đã có", "OK", "restart: unless-stopped"),
        ("6", "Backup", "Backup database (manual)",
         "pg_dump full database",
         "docker exec openwebui-postgres pg_dump ...", "Đã có", "Chưa test", "Cần chạy thủ công"),
        ("7", "Network & Firewall", "Internal Docker network",
         "Các container giao tiếp qua mạng Docker riêng, không expose port nội bộ",
         "", "Đã có", "OK", "openwebui-network"),
        ("8", "", "Firewall port 3000, 5000",
         "Đã mở firewall cho truy cập từ bên ngoài",
         "", "Đã có", "OK", ""),
    ]),
    ("IX. TÍNH NĂNG DÀNH CHO NGƯỜI DÙNG", [
        ("1", "Giao diện", "Giao diện web responsive",
         "Truy cập qua trình duyệt, hỗ trợ desktop và mobile",
         "", "Đã có", "OK", ""),
        ("2", "", "Dark mode / Light mode",
         "Chuyển đổi giao diện sáng/tối",
         "Settings → Theme", "Đã có", "OK", ""),
        ("3", "", "Đa ngôn ngữ (tiếng Việt, Anh, ...)",
         "Giao diện hỗ trợ nhiều ngôn ngữ",
         "Settings → Language", "Đã có", "OK", ""),
        ("4", "Cá nhân hoá", "Chọn model mặc định",
         "User chọn model ưa thích làm mặc định cho mọi chat mới",
         "Settings → Default Model", "Đã có", "OK", ""),
        ("5", "", "System prompt cá nhân",
         "User tự đặt system prompt riêng để tuỳ chỉnh phong cách AI",
         "Settings → System Prompt", "Đã có", "OK", ""),
        ("6", "", "Memory – AI nhớ thông tin user",
         "AI lưu thông tin cá nhân user để cá nhân hoá câu trả lời",
         "Nhớ rằng tôi là developer Python", "Đã có", "OK", "Table: memory"),
        ("7", "Tìm kiếm", "Tìm kiếm trong lịch sử chat",
         "Tìm theo keyword trong tiêu đề hoặc nội dung",
         "Thanh search trên sidebar", "Đã có", "OK", ""),
        ("8", "Xuất dữ liệu", "Export chat history",
         "Xuất lịch sử chat thành file", "", "Đã có", "OK", ""),
        ("9", "", "Xuất Excel/PDF/DOCX",
         "Sử dụng custom tools để xuất dữ liệu có format chuyên nghiệp",
         "Action → Xuất Excel/PDF/DOCX", "Đã có", "OK", ""),
        ("10", "Prompt", "Saved Prompts (mẫu prompt)",
         "Lưu prompt hay dùng để tái sử dụng nhanh",
         "Workspace → Prompts", "Đã có", "Chưa test", "Hạ tầng sẵn"),
        ("11", "", "Prompt suggestions",
         "Gợi ý câu hỏi khi mở chat mới",
         "", "Đã có", "OK", "Admin cấu hình"),
        ("12", "Channels", "Kênh chat nhóm",
         "Tạo kênh chat nhóm, mời thành viên, chia sẻ thông tin",
         "", "Đã có", "Chưa test", "Hạ tầng sẵn"),
    ]),
    ("X. TÍNH NĂNG DÀNH CHO ADMIN", [
        ("1", "Quản lý User", "Danh sách user",
         "Xem tất cả users, role, trạng thái, thời gian hoạt động",
         "Admin Panel → Users", "Đã có", "OK", ""),
        ("2", "", "Duyệt user mới",
         "Approve/Reject user pending", "", "Đã có", "OK", ""),
        ("3", "", "Đổi role user",
         "Thăng/hạ quyền (Admin ↔ User)", "", "Đã có", "OK", ""),
        ("4", "", "Xoá user",
         "Xoá tài khoản user và dữ liệu liên quan", "", "Đã có", "OK", ""),
        ("5", "Quản lý Model", "Bật/tắt model",
         "Enable/disable model cho users", "Admin → Models", "Đã có", "OK", ""),
        ("6", "", "Cấu hình model params",
         "Set temperature, top_p, max_tokens mặc định", "", "Đã có", "OK", ""),
        ("7", "", "Gán Knowledge vào model",
         "Model tự động dùng Knowledge cụ thể", "", "Đã có", "OK", ""),
        ("8", "Quản lý Knowledge", "Xem/xoá tất cả Knowledge",
         "Liệt kê toàn bộ Knowledge Collections, xoá + cascade embeddings",
         "Admin → Knowledge", "Đã có", "OK", ""),
        ("9", "Cấu hình RAG", "Điều chỉnh chunk size, file limit, embedding model",
         "Thay đổi qua biến môi trường: CHUNK_SIZE, RAG_FILE_MAX_SIZE, RAG_EMBEDDING_MODEL",
         "", "Đã có", "OK", "Cần restart container"),
        ("10", "Cấu hình hệ thống", "WebUI settings",
         "Signup, default model, banner, announcements",
         "Admin → Settings → General", "Đã có", "OK", ""),
        ("11", "", "Connections",
         "Cấu hình kết nối đến LLM providers",
         "Admin → Settings → Connections", "Đã có", "OK", ""),
        ("12", "Giám sát", "Request logs chi tiết",
         "Xem log từng API request: user, model, tokens, cost",
         "logs/middleware.requests.log", "Đã có", "OK", ""),
        ("13", "", "Application logs",
         "Xem lỗi, warning, events hệ thống",
         "logs/middleware.log", "Đã có", "OK", ""),
        ("14", "", "Cost dashboard",
         "Báo cáo chi phí theo user/model/thời gian",
         "http://<server>:5000/dashboard", "Đã có", "OK", ""),
    ]),
    ("XI. TÍNH NĂNG TRONG KẾ HOẠCH PHÁT TRIỂN", [
        ("1", "Quản lý nhóm", "Tạo nhóm theo phòng ban",
         "Phân quyền và quota theo nhóm. Framework có sẵn (table group, group_member)",
         "", "Chưa có", "", "Kế hoạch ngắn hạn"),
        ("2", "Backup", "Backup tự động database",
         "Cron job pg_dump hàng ngày/tuần",
         "", "Chưa có", "", "Kế hoạch ngắn hạn"),
        ("3", "Monitoring", "Uptime monitoring + alerting",
         "Prometheus + Grafana hoặc tương đương",
         "", "Chưa có", "", "Kế hoạch trung hạn"),
        ("4", "SSO/LDAP", "Đăng nhập bằng AD nội bộ",
         "Tích hợp Active Directory/LDAP. Open WebUI hỗ trợ sẵn",
         "", "Chưa có", "", "Kế hoạch trung hạn"),
        ("5", "API integration", "Tích hợp DMS, ERP nội bộ",
         "Kết nối Open WebUI với hệ thống nội bộ qua API",
         "", "Chưa có", "", "Kế hoạch trung hạn"),
        ("6", "Scheduled reports", "Báo cáo chi phí tự động",
         "Email/Zalo gửi báo cáo hàng tuần",
         "", "Chưa có", "", "Kế hoạch trung hạn"),
        ("7", "Fine-tuning", "Huấn luyện model riêng trên data nội bộ",
         "Train model custom. Cần GPU",
         "", "Chưa có", "", "Kế hoạch dài hạn"),
        ("8", "On-premise LLM", "Chạy AI local (Llama, Mistral)",
         "Không phụ thuộc API bên ngoài. Cần GPU mạnh",
         "", "Chưa có", "", "Kế hoạch dài hạn"),
    ]),
]


# ============================================================
# STYLES
# ============================================================
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E79")
MODULE_FONT = Font(bold=True, size=12, color="1F4E79")
MODULE_FILL = PatternFill("solid", fgColor="D6E4F0")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
CELL_FONT = Font(size=10)
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

STATUS_COLORS = {
    "Đã có": PatternFill("solid", fgColor="E2EFDA"),
    "Chưa có": PatternFill("solid", fgColor="FCE4D6"),
    "Hạ tầng có": PatternFill("solid", fgColor="FFF2CC"),
    "Đang phát triển": PatternFill("solid", fgColor="FFF2CC"),
}
TEST_COLORS = {
    "OK": PatternFill("solid", fgColor="E2EFDA"),
    "Đang lỗi": PatternFill("solid", fgColor="FCE4D6"),
    "Chưa test": PatternFill("solid", fgColor="FFF2CC"),
}

COL_WIDTHS = [6, 22, 22, 38, 48, 35, 16, 14, 22]


def build():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Remove old sheet if exists, create new
    sheet_name = "list tính năng"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 1)  # index 1

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title row
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1, value=TITLE)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 35

    # Blank row
    row = 2

    # Header row
    row = 3
    for i, h in enumerate(HEADER_ROW, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = "A4"

    row = 4
    for module_name, items in MODULES:
        # Module header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value=module_name)
        c.font = MODULE_FONT
        c.fill = MODULE_FILL
        c.alignment = Alignment(vertical="center")
        c.border = THIN_BORDER
        ws.row_dimensions[row].height = 26
        row += 1

        # Data rows
        for item in items:
            stt, nhom, tinh_nang, huong_dan, vi_du, trang_thai, ket_qua, ghi_chu = item
            values = [stt, "", nhom, tinh_nang, huong_dan, vi_du, trang_thai, ket_qua, ghi_chu]
            for i, v in enumerate(values, 1):
                c = ws.cell(row=row, column=i, value=v)
                c.font = CELL_FONT
                c.alignment = WRAP_ALIGN if i >= 3 else CENTER_ALIGN
                c.border = THIN_BORDER

                # Color coding
                if i == 7 and v in STATUS_COLORS:
                    c.fill = STATUS_COLORS[v]
                    c.alignment = CENTER_ALIGN
                if i == 8 and v in TEST_COLORS:
                    c.fill = TEST_COLORS[v]
                    c.alignment = CENTER_ALIGN

            row += 1

        # Blank separator
        row += 1

    # Summary row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    total_done = sum(len([r for r in items if r[5] == "Đã có" or r[5] == "Hạ tầng có"]) for _, items in MODULES)
    total_planned = sum(len([r for r in items if r[5] == "Chưa có"]) for _, items in MODULES)
    c = ws.cell(row=row, column=1, value=f"TỔNG KẾT: {total_done} tính năng đã hoạt động | {total_planned} tính năng trong kế hoạch phát triển")
    c.font = Font(bold=True, size=11, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    wb.save(EXCEL_PATH)
    print(f"OK: Saved to {EXCEL_PATH}")
    print(f"  Sheet: {sheet_name}")
    print(f"  Rows: {row}")
    print(f"  Features done: {total_done}, planned: {total_planned}")


if __name__ == "__main__":
    build()
