"""
Export Report API - generates a downloadable Excel (multi-sheet) or CSV
export of aggregated dashboard data (MW DB + OW DB) for a given time range.
"""

import io
import csv
import uuid
import datetime as dt
from collections import defaultdict
from typing import Optional

from fastapi import Request, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.db import db_conn, db_ow_conn, get_conn, put_conn, fetch_final_audit_entries
from utils.auth_guard import require_admin_or_session


# ─── Time range parsing (same rules as summary_v2.get_summary_v2) ────────

def _parse_time_range(start: Optional[str], end: Optional[str]):
    now_utc = dt.datetime.now(tz=dt.timezone.utc)

    if start and end:
        try:
            start_normalized = start.replace('Z', '+00:00') if start.endswith('Z') else start
            end_normalized = end.replace('Z', '+00:00') if end.endswith('Z') else end

            cutoff = dt.datetime.fromisoformat(start_normalized)
            end_time = dt.datetime.fromisoformat(end_normalized)

            if cutoff.tzinfo is None:
                cutoff = cutoff.replace(tzinfo=dt.timezone.utc)
            if end_time.tzinfo is None:
                end_time = end_time.replace(tzinfo=dt.timezone.utc)

            if cutoff >= end_time:
                raise HTTPException(400, "Start time must be before end time")
        except ValueError as e:
            raise HTTPException(400, f"Invalid datetime format: {e}")
    else:
        cutoff = now_utc - dt.timedelta(minutes=60)
        end_time = now_utc

    return cutoff, end_time


# ─── Data collection layer ────────────────────────────────────────────────

def _collect_summary(request: Request, cutoff, end_time) -> dict:
    """Reuse summary_v2's aggregation (same DB source, same totals shape)."""
    from api.summary_v2 import get_summary_v2
    try:
        data = get_summary_v2(request, start=cutoff.isoformat(), end=end_time.isoformat())
        if isinstance(data, dict) and "totals" in data:
            return data["totals"]
    except Exception:
        pass
    return {}


def _collect_top_users(cutoff, end_time) -> list:
    """Per-user breakdown (requests/cost/tokens/top model) with OW display names."""
    user_stats = defaultdict(lambda: {"requests": 0, "cost_usd": 0.0, "tokens": 0, "models": defaultdict(int)})
    try:
        for entry in fetch_final_audit_entries(cutoff, end_time):
            is_final = entry["status"] in ('ok', 'reconciled')
            stats = user_stats[entry["user_id"] or "unknown"]
            stats["requests"] += 1
            stats["cost_usd"] += entry["cost_usd"] if is_final else 0.0
            stats["tokens"] += entry["tokens_total"] if is_final else 0
            stats["models"][entry["model"] or "unknown"] += 1
    except Exception:
        pass

    display_names = {}
    try:
        with db_ow_conn() as conn:
            cur = conn.cursor()
            cur.execute('SELECT email, name FROM "user"')
            for email, name in cur.fetchall():
                display_names[email] = name or email
            cur.close()
    except Exception:
        pass

    rows = []
    for user_id, stats in user_stats.items():
        top_model = max(stats["models"].items(), key=lambda x: x[1])[0] if stats["models"] else "unknown"
        rows.append({
            "user_id": user_id,
            "display_name": display_names.get(user_id, user_id),
            "requests": stats["requests"],
            "cost_usd": round(stats["cost_usd"], 6),
            "tokens": stats["tokens"],
            "top_model": top_model,
        })
    rows.sort(key=lambda x: x["cost_usd"], reverse=True)
    return rows


def _collect_top_models(cutoff, end_time) -> list:
    """Model breakdown (requests/cost/tokens)."""
    model_stats = defaultdict(lambda: {"requests": 0, "cost_usd": 0.0, "tokens": 0})
    try:
        for entry in fetch_final_audit_entries(cutoff, end_time):
            is_final = entry["status"] in ('ok', 'reconciled')
            stats = model_stats[entry["model"] or "unknown"]
            stats["requests"] += 1
            stats["cost_usd"] += entry["cost_usd"] if is_final else 0.0
            stats["tokens"] += entry["tokens_total"] if is_final else 0
    except Exception:
        pass

    rows = [
        {"model": m, "requests": s["requests"], "cost_usd": round(s["cost_usd"], 6), "tokens": s["tokens"]}
        for m, s in model_stats.items()
    ]
    rows.sort(key=lambda x: x["cost_usd"], reverse=True)
    return rows


def _minutes_between(cutoff, end_time) -> int:
    """Integer minute span, for functions whose `minutes` param is a plain int
    default (not Optional[int] = None) and may be dereferenced even when
    start/end are also given — must not be left as FastAPI's raw Query() sentinel."""
    return max(1, int((end_time - cutoff).total_seconds() // 60))


def _collect_groups(request: Request, cutoff, end_time) -> list:
    """Reuse group_analytics's aggregation. Empty list means 'unavailable' to the caller."""
    from api.group_analytics import get_group_analytics
    try:
        data = get_group_analytics(
            request, minutes=_minutes_between(cutoff, end_time),
            start=cutoff.isoformat(), end=end_time.isoformat()
        )
        if isinstance(data, dict):
            return data.get("groups", [])
    except Exception:
        pass
    return []


def _collect_chat_analytics(request: Request, cutoff, end_time) -> dict:
    """Reuse analytics.get_chat_analytics's aggregation (OW chat/message + MW cost data)."""
    from api.analytics import get_chat_analytics
    try:
        return get_chat_analytics(
            request, minutes=_minutes_between(cutoff, end_time),
            start=cutoff.isoformat(), end=end_time.isoformat()
        )
    except Exception:
        return {}


def _collect_satisfaction(request: Request, cutoff, end_time) -> dict:
    """Reuse analytics.get_satisfaction_analytics's aggregation (OW feedback)."""
    from api.analytics import get_satisfaction_analytics
    try:
        return get_satisfaction_analytics(
            request, minutes=_minutes_between(cutoff, end_time),
            start=cutoff.isoformat(), end=end_time.isoformat()
        )
    except Exception:
        return {}


_AUDIT_COLUMNS_SQL = """
    SELECT ts, rid, user_id, endpoint, model, purpose, status,
           status_code, latency_ms, tokens_in, tokens_out, tokens_total,
           cost_usd, image_count, tts_chars, stt_seconds, video_count,
           error_type, error_message
    FROM mw_audit_log
    WHERE ts >= %s AND ts <= %s
    ORDER BY ts DESC
"""


def _collect_audit_log(cutoff, end_time, limit: int = 50000) -> list:
    """Raw audit log records, most recent first, capped at `limit` rows."""
    rows = []
    try:
        with db_conn() as conn:
            cur = conn.cursor()
            cur.execute(_AUDIT_COLUMNS_SQL + " LIMIT %s", (cutoff, end_time, limit))
            for r in cur.fetchall():
                rows.append({
                    "ts": r[0].isoformat() if r[0] else "",
                    "rid": r[1] or "", "user_id": r[2] or "",
                    "endpoint": r[3] or "", "model": r[4] or "",
                    "purpose": r[5] or "", "status": r[6] or "",
                    "status_code": r[7], "latency_ms": r[8],
                    "tokens_in": r[9] or 0, "tokens_out": r[10] or 0,
                    "tokens_total": r[11] or 0, "cost_usd": float(r[12] or 0),
                    "image_count": r[13], "tts_chars": r[14],
                    "stt_seconds": r[15], "video_count": r[16],
                    "error_type": r[17] or "", "error_message": r[18] or "",
                })
            cur.close()
    except Exception:
        pass
    return rows


# ─── Excel generation ─────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _autosize_columns(ws, ncols):
    widths = [0] * ncols
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ncols):
        for idx, cell in enumerate(row):
            if cell.value is not None:
                length = len(str(cell.value))
                if length > widths[idx]:
                    widths[idx] = length
    for idx in range(ncols):
        letter = get_column_letter(idx + 1)
        ws.column_dimensions[letter].width = min(max(widths[idx] + 2, 10), 50)


def _write_table(ws, headers, rows, start_row: int = 1) -> int:
    """Write a header row + data rows starting at start_row.
    Applies bold/colored header, auto-width, freeze panes and AutoFilter.
    Returns the last written row index."""
    ncols = len(headers)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    r = start_row + 1
    for row in rows:
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)
        r += 1

    last_row = max(r - 1, start_row)
    last_col_letter = get_column_letter(ncols)
    ws.auto_filter.ref = f"A{start_row}:{last_col_letter}{last_row}"
    ws.freeze_panes = f"A{start_row + 1}"
    _autosize_columns(ws, ncols)
    return last_row


def _generate_xlsx(request: Request, cutoff, end_time) -> bytes:
    wb = Workbook()

    # Sheet 1: Tổng quan
    ws = wb.active
    ws.title = "Tổng quan"
    summary = _collect_summary(request, cutoff, end_time)
    kv_rows = [
        ("Khoảng thời gian", f"{cutoff.isoformat()} → {end_time.isoformat()}"),
        ("Tổng số requests", summary.get("requests_total", 0)),
        ("Tổng chi phí (USD)", round(summary.get("cost_total_usd", 0) or 0, 6)),
        ("Tổng tokens", summary.get("tokens_total", 0)),
        ("Tỷ lệ lỗi (%)", summary.get("error_rate_percent", 0)),
        ("P95 Latency (ms)", summary.get("p95_latency_ms") or "-"),
        ("Chat calls", summary.get("chat_calls", 0)),
        ("Image calls", summary.get("image_calls", 0)),
        ("Audio calls", summary.get("audio_calls", 0)),
        ("Video calls", summary.get("video_calls", 0)),
    ]
    _write_table(ws, ["Chỉ số", "Giá trị"], kv_rows)

    # Sheet 2: Top Users
    ws2 = wb.create_sheet("Top Users")
    top_users = _collect_top_users(cutoff, end_time)
    _write_table(
        ws2, ["User ID", "Display Name", "Requests", "Cost (USD)", "Tokens", "Top Model"],
        [(u["user_id"], u["display_name"], u["requests"], u["cost_usd"], u["tokens"], u["top_model"]) for u in top_users]
    )

    # Sheet 3: Top Models
    ws3 = wb.create_sheet("Top Models")
    top_models = _collect_top_models(cutoff, end_time)
    _write_table(
        ws3, ["Model", "Requests", "Cost (USD)", "Tokens"],
        [(m["model"], m["requests"], m["cost_usd"], m["tokens"]) for m in top_models]
    )

    # Sheet 4: Phòng ban
    ws4 = wb.create_sheet("Phòng ban")
    groups = _collect_groups(request, cutoff, end_time)
    if groups:
        rows4 = []
        for g in groups:
            model_prefs = g.get("model_preferences") or []
            top_model = model_prefs[0]["model"] if model_prefs else "-"
            rows4.append((
                g.get("group_name", "Uncategorized"), g.get("total_requests", 0),
                g.get("total_cost", 0), g.get("total_tokens", 0),
                g.get("avg_latency_ms", 0), top_model
            ))
        _write_table(ws4, ["Group Name", "Requests", "Cost (USD)", "Tokens", "Avg Latency (ms)", "Top Model"], rows4)
    else:
        ws4.cell(row=1, column=1, value="Dữ liệu nhóm không khả dụng")

    # Sheet 5: Chat Analytics
    ws5 = wb.create_sheet("Chat Analytics")
    chat = _collect_chat_analytics(request, cutoff, end_time)
    totals5 = chat.get("totals", {})
    ws5.cell(row=1, column=1, value="Tổng số chats")
    ws5.cell(row=1, column=2, value=totals5.get("chats", 0))
    ws5.cell(row=2, column=1, value="Tổng số messages")
    ws5.cell(row=2, column=2, value=totals5.get("requests", 0))
    ws5.cell(row=3, column=1, value="Active users")
    ws5.cell(row=3, column=2, value=totals5.get("active_users", 0))
    leaderboard = chat.get("leaderboard", [])
    _write_table(
        ws5, ["User", "Display Name", "Chat Count", "Request Count", "Cost (USD)", "Top Model"],
        [(u["user_id"], u["display_name"], u["chat_count"], u["request_count"], u["cost_usd"], u["top_model"]) for u in leaderboard],
        start_row=5
    )

    # Sheet 6: Satisfaction
    ws6 = wb.create_sheet("Satisfaction")
    sat = _collect_satisfaction(request, cutoff, end_time)
    totals6 = sat.get("totals", {})
    ws6.cell(row=1, column=1, value="Tổng feedback")
    ws6.cell(row=1, column=2, value=totals6.get("total", 0))
    ws6.cell(row=2, column=1, value="Positive")
    ws6.cell(row=2, column=2, value=totals6.get("positive", 0))
    ws6.cell(row=3, column=1, value="Negative")
    ws6.cell(row=3, column=2, value=totals6.get("negative", 0))
    ws6.cell(row=4, column=1, value="CSAT %")
    ws6.cell(row=4, column=2, value=totals6.get("csat_percent", 0))
    model_leaderboard = sat.get("model_leaderboard", [])
    _write_table(
        ws6, ["Model", "Positive", "Negative", "Total", "CSAT %"],
        [(m["model_id"], m["positive"], m["negative"], m["total"], m["csat_percent"]) for m in model_leaderboard],
        start_row=6
    )

    # Sheet 7: Audit Log
    ws7 = wb.create_sheet("Audit Log")
    audit_rows = _collect_audit_log(cutoff, end_time, limit=50000)
    _write_table(
        ws7, ["Timestamp", "Request ID", "User ID", "Endpoint", "Model", "Status",
              "Latency (ms)", "Tokens In", "Tokens Out", "Cost (USD)", "Error Type"],
        [(r["ts"], r["rid"], r["user_id"], r["endpoint"], r["model"], r["status"],
          r["latency_ms"], r["tokens_in"], r["tokens_out"], r["cost_usd"], r["error_type"])
         for r in audit_rows]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── CSV streaming ────────────────────────────────────────────────────────

_CSV_HEADER = [
    "Timestamp", "Request ID", "User ID", "Endpoint", "Model", "Purpose", "Status",
    "Status Code", "Latency (ms)", "Tokens In", "Tokens Out", "Tokens Total",
    "Cost (USD)", "Image Count", "TTS Chars", "STT Seconds", "Video Count",
    "Error Type", "Error Message"
]


def _generate_csv_stream(cutoff, end_time):
    """Generator yielding CSV text chunks for the audit log, using a server-side
    cursor + fetchmany() so the full result set is never buffered in memory."""
    yield '﻿'  # UTF-8 BOM for Excel compatibility

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_CSV_HEADER)
    yield buf.getvalue()

    conn = get_conn()
    try:
        cur = conn.cursor(name=f"export_audit_{uuid.uuid4().hex}")
        cur.itersize = 1000
        cur.execute(_AUDIT_COLUMNS_SQL, (cutoff, end_time))
        while True:
            rows = cur.fetchmany(1000)
            if not rows:
                break
            buf = io.StringIO()
            writer = csv.writer(buf)
            for r in rows:
                writer.writerow([
                    r[0].isoformat() if r[0] else "", r[1] or "", r[2] or "",
                    r[3] or "", r[4] or "", r[5] or "", r[6] or "",
                    r[7], r[8], r[9] or 0, r[10] or 0, r[11] or 0,
                    float(r[12] or 0), r[13], r[14], r[15], r[16],
                    r[17] or "", r[18] or "",
                ])
            yield buf.getvalue()
        cur.close()
        conn.commit()
    except Exception:
        conn.rollback()
    finally:
        put_conn(conn)


# ─── API endpoint ──────────────────────────────────────────────────────────

def export_report(
    request: Request,
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    start: Optional[str] = Query(None),
    end: Optional[str] = Query(None),
):
    """
    GET /v1/_mw/export/report?format=xlsx|csv&start=...&end=...
    Admin-only. Generates a downloadable multi-sheet Excel report or a
    streaming CSV of the full audit log for the given time range.
    """
    require_admin_or_session(request)
    cutoff, end_time = _parse_time_range(start, end)
    date_tag = f"{cutoff.strftime('%Y%m%d')}_to_{end_time.strftime('%Y%m%d')}"

    if format == "csv":
        from core.db import _pool
        if _pool is None:
            raise HTTPException(503, "Database unavailable for CSV export")
        filename = f"LLM_AuditLog_{date_tag}.csv"
        return StreamingResponse(
            _generate_csv_stream(cutoff, end_time),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    filename = f"LLM_Report_{date_tag}.xlsx"
    content = _generate_xlsx(request, cutoff, end_time)
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
