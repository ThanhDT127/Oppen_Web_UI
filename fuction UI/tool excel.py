"""
title: Xuất Excel (Wizard dài + chuẩn hoá theo context hội thoại)
author: Thanh + ChatGPT
version: 3.1.0
required_open_webui_version: 0.6.0
requirements: openpyxl,requests
"""

import asyncio
import base64
import datetime
import io
import logging
import re
from typing import Any, Dict, List, Optional

import requests
from pydantic import BaseModel, Field

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class Action:
    class Valves(BaseModel):
        # ---- UX ----
        show_status: bool = Field(
            default=True, description="Hiển thị status/notification"
        )
        show_confirmation: bool = Field(
            default=False, description="Hỏi xác nhận trước khi chạy"
        )
        min_wizard_seconds: float = Field(
            default=1.5, description="Giữ wizard tối thiểu (giây)"
        )

        # ---- OpenWebUI internal call ----
        openwebui_base_url: str = Field(
            default="",
            description="Base URL Open WebUI (VD: http://localhost:3000). Để trống thì cố suy ra từ request.",
        )
        openwebui_chat_completions_path: str = Field(
            default="/api/chat/completions", description="Path chat completions"
        )
        timeout_sec: int = Field(
            default=120, description="Timeout gọi /api/chat/completions"
        )

        # ---- Normalize scope ----
        # assistant_only: chỉ dùng câu trả lời bot cuối cùng để chuẩn hoá
        # conversation_window: dùng N messages gần nhất để chuẩn hoá theo đúng yêu cầu user
        normalize_mode: str = Field(
            default="conversation_window",
            description="assistant_only | conversation_window",
        )
        context_window_messages: int = Field(
            default=14,
            description="Số messages gần nhất đưa vào normalize (khi conversation_window)",
        )
        drop_file_like_content: bool = Field(
            default=True,
            description="Loại các đoạn giống file/base64 khỏi context (giảm token)",
        )

        # ---- Normalize prompt ----
        normalize_with_llm: bool = Field(
            default=True, description="Gọi LLM để chuẩn hoá về Markdown table"
        )
        max_rows_hint: int = Field(
            default=250, description="Gợi ý giới hạn số dòng dữ liệu"
        )
        normalize_instructions_vi: str = Field(
            default=(
                "Bạn là trợ lý chuẩn hoá dữ liệu.\n"
                "Nhiệm vụ: tạo ĐÚNG 1 bảng Markdown duy nhất từ nội dung hội thoại.\n\n"
                "YÊU CẦU BẮT BUỘC:\n"
                "1) Chỉ trả về bảng Markdown, KHÔNG thêm bất kỳ văn bản nào trước/sau, KHÔNG bọc ```.\n"
                "2) Dòng 1 là header, dòng 2 là separator |---|.\n"
                "3) Header ngắn gọn, rõ nghĩa, theo đúng yêu cầu người dùng.\n"
                "4) Không xuống dòng trong ô.\n"
                "5) Nếu thiếu dữ liệu, để trống ô.\n"
                "6) Không tạo nhiều bảng.\n"
                "7) Nếu người dùng yêu cầu lọc/cột/số dòng/format thì phải tuân thủ.\n"
            ),
            description="System prompt chuẩn hoá bảng (VI)",
        )

        # ---- Excel output ----
        sheet_name: str = Field(default="Extracted", description="Tên sheet")
        base_filename: str = Field(
            default="trich_xuat.xlsx", description="Tên file (có timestamp)"
        )
        title: str = Field(
            default="BẢNG TRÍCH XUẤT", description="Tiêu đề trên đầu file"
        )
        table_index: int = Field(
            default=0, description="Nếu output có nhiều bảng, chọn bảng N (0-based)"
        )
        table_start_row: int = Field(
            default=5, description="Dòng bắt đầu bảng (header)"
        )
        max_col_width: int = Field(default=55, description="Độ rộng cột tối đa")

    def __init__(self):
        self.valves = self.Valves()

    # =========================================================
    # UI MODAL
    # =========================================================

    async def _ui_modal_open(self, __event_call__=None):
        if not __event_call__:
            return
        js = """
(() => {
  const ID = "owui_export_excel_modal";
  if (document.getElementById(ID)) return;

  const overlay = document.createElement("div");
  overlay.id = ID;
  overlay.style.cssText = `
    position: fixed; inset: 0; z-index: 999999;
    background: rgba(0,0,0,0.45);
    display: flex; align-items: center; justify-content: center;
    font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial;
  `;

  overlay.innerHTML = `
    <div style="
      width: min(620px, 94vw);
      background: #fff; border-radius: 14px;
      box-shadow: 0 16px 60px rgba(0,0,0,.25);
      padding: 18px 18px 14px 18px;
    ">
      <div style="display:flex; gap:12px; align-items:flex-start;">
        <div id="owui_modal_spinner" style="
          width: 34px; height: 34px; border-radius: 999px;
          border: 3px solid #d1d5db; border-top-color: #2563eb;
          animation: spin 0.9s linear infinite; margin-top: 2px;
        "></div>

        <div style="flex:1">
          <div style="display:flex; justify-content:space-between; gap:12px; align-items:flex-start;">
            <div>
              <div style="font-size:16px; font-weight:700; color:#111827;" id="owui_modal_title">Đang xuất Excel</div>
              <div style="font-size:13px; color:#4b5563; margin-top:4px;" id="owui_modal_subtitle">Khởi động tiến trình...</div>
            </div>
            <div style="font-size:12px; color:#6b7280; margin-top:2px;">
              <span id="owui_modal_timer">0</span>s
            </div>
          </div>

          <div style="margin-top:12px; padding:10px 10px; background:#f9fafb; border:1px solid #e5e7eb; border-radius:10px;">
            <div style="font-size:12px; color:#374151; font-weight:600; margin-bottom:6px;">Tiến trình</div>
            <ol style="margin:0; padding-left:18px; font-size:12.5px; color:#111827; line-height:1.65" id="owui_modal_steps">
              <li>Khởi tạo tiến trình</li>
            </ol>
          </div>

          <div style="display:flex; justify-content:flex-end; gap:10px; margin-top:14px;">
            <button id="owui_modal_close" disabled
              style="opacity:.6; cursor:not-allowed; padding:8px 12px; border-radius:10px; border:1px solid #e5e7eb; background:#fff;">
              Đóng
            </button>
          </div>

          <div style="font-size:11px; color:#6b7280; margin-top:10px;">
            Thao tác có thể mất 10–60 giây tuỳ độ dài hội thoại và tốc độ model.
          </div>
        </div>
      </div>
    </div>
  `;

  const style = document.createElement("style");
  style.textContent = `@keyframes spin { to { transform: rotate(360deg); } }`;
  overlay.appendChild(style);

  let sec = 0;
  const timerEl = overlay.querySelector("#owui_modal_timer");
  const t = setInterval(() => {
    sec += 1;
    if (timerEl) timerEl.textContent = String(sec);
  }, 1000);
  overlay.dataset.timer = String(t);

  document.body.appendChild(overlay);
})();
"""
        await __event_call__({"type": "execute", "data": {"code": js}})

    async def _ui_modal_update(
        self,
        __event_call__=None,
        title: Optional[str] = None,
        subtitle: Optional[str] = None,
        steps: Optional[List[str]] = None,
        done_ok: Optional[bool] = None,
        error_text: Optional[str] = None,
    ):
        if not __event_call__:
            return

        def _js_str(s: str) -> str:
            return s.replace("\\", "\\\\").replace("`", "\\`").replace("$", "\\$")

        steps_js = "null"
        if isinstance(steps, list):
            safe = [_js_str(x) for x in steps]
            steps_js = "[" + ",".join([f"`{x}`" for x in safe]) + "]"

        title_js = f"`{_js_str(title)}`" if isinstance(title, str) else "null"
        subtitle_js = f"`{_js_str(subtitle)}`" if isinstance(subtitle, str) else "null"
        error_js = f"`{_js_str(error_text)}`" if isinstance(error_text, str) else "null"
        done_js = (
            "true" if done_ok is True else ("false" if done_ok is False else "null")
        )

        js = f"""
(() => {{
  const overlay = document.getElementById("owui_export_excel_modal");
  if (!overlay) return;

  const t = overlay.querySelector("#owui_modal_title");
  const st = overlay.querySelector("#owui_modal_subtitle");
  const stepsEl = overlay.querySelector("#owui_modal_steps");
  const closeBtn = overlay.querySelector("#owui_modal_close");
  const spinner = overlay.querySelector("#owui_modal_spinner");

  const title = {title_js};
  const subtitle = {subtitle_js};
  const steps = {steps_js};
  const done = {done_js};
  const err = {error_js};

  if (title) t.textContent = title;
  if (subtitle) st.textContent = subtitle;

  if (Array.isArray(steps) && stepsEl) {{
    stepsEl.innerHTML = steps.map(s => `<li>${{s}}</li>`).join("");
  }}

  if (done !== null) {{
    const tid = overlay.dataset.timer;
    if (tid) {{
      try {{ clearInterval(Number(tid)); }} catch (e) {{}}
    }}

    closeBtn.disabled = false;
    closeBtn.style.opacity = "1";
    closeBtn.style.cursor = "pointer";
    closeBtn.onclick = () => overlay.remove();

    if (spinner) {{
      spinner.style.animation = "none";
      spinner.style.borderTopColor = done ? "#16a34a" : "#dc2626";
      spinner.style.borderColor = done ? "#86efac" : "#fecaca";
      spinner.style.background = done ? "#dcfce7" : "#fee2e2";
    }}

    if (done === true) {{
      t.textContent = "Xuất Excel thành công";
      st.textContent = "File đang được tải xuống. Bạn có thể đóng cửa sổ này.";
    }} else {{
      t.textContent = "Xuất Excel thất bại";
      st.textContent = err ? err : "Có lỗi xảy ra. Vui lòng thử lại hoặc liên hệ quản trị.";
    }}
  }}
}})();
"""
        await __event_call__({"type": "execute", "data": {"code": js}})

    # =========================================================
    # Context helpers
    # =========================================================

    def _get_current_model_id(self, body: dict, __model__=None) -> Optional[str]:
        m = body.get("model")
        if isinstance(m, str) and m.strip():
            return m.strip()
        if isinstance(__model__, dict):
            for k in ("id", "model", "name"):
                v = __model__.get(k)
                if isinstance(v, str) and v.strip():
                    return v.strip()
        return None

    def _get_messages_window(self, body: dict) -> List[Dict[str, Any]]:
        msgs = body.get("messages", [])
        if not isinstance(msgs, list):
            return []
        n = (
            int(self.valves.context_window_messages)
            if self.valves.context_window_messages
            else 0
        )
        if n <= 0:
            return msgs
        return msgs[-n:]

    def _sanitize_message_content(self, content: Any) -> str:
        """
        Giảm rủi ro nổ token:
        - Nếu content là list/blocks (multimodal) -> cố chuyển sang text ngắn.
        - Nếu có base64/data url -> cắt bỏ.
        """
        if content is None:
            return ""

        if isinstance(content, str):
            s = content
        else:
            # multimodal content sometimes is list of dicts
            try:
                s = str(content)
            except Exception:
                s = ""

        if self.valves.drop_file_like_content:
            # remove base64-like huge strings
            if "data:" in s and "base64" in s:
                s = re.sub(
                    r"data:[^\\s]+;base64,[A-Za-z0-9+/=]+", "[DATA_URL_REMOVED]", s
                )
            # remove long base64 blocks
            s = re.sub(r"[A-Za-z0-9+/=]{1500,}", "[BASE64_REMOVED]", s)

        # hard cap per message to keep normalize stable
        if len(s) > 6000:
            s = s[:6000] + "…[TRUNCATED]"
        return s

    def _build_normalize_user_prompt(self, body: dict) -> str:
        mode = (self.valves.normalize_mode or "conversation_window").strip()

        if mode == "assistant_only":
            # Only last assistant content
            last_assistant = ""
            for m in reversed(body.get("messages", [])):
                if isinstance(m, dict) and m.get("role") == "assistant":
                    last_assistant = self._sanitize_message_content(m.get("content"))
                    if last_assistant.strip():
                        break
            return (
                "Hãy tạo bảng Markdown từ nội dung TRẢ LỜI CỦA BOT dưới đây.\n\n"
                f"BOT:\n{last_assistant}"
            )

        # conversation_window (default)
        window = self._get_messages_window(body)
        lines = []
        for m in window:
            if not isinstance(m, dict):
                continue
            role = m.get("role", "")
            c = self._sanitize_message_content(m.get("content"))
            if not c.strip():
                continue
            if role == "system":
                # thường bỏ system của hội thoại để tránh xung đột; nhưng nếu có thì keep ngắn
                lines.append(f"[SYSTEM]\n{c}")
            elif role == "user":
                lines.append(f"[USER]\n{c}")
            elif role == "assistant":
                lines.append(f"[ASSISTANT]\n{c}")
            else:
                lines.append(f"[{role.upper()}]\n{c}")

        # add explicit constraint
        return (
            "Dưới đây là đoạn hội thoại gần nhất. Hãy dựa vào YÊU CẦU CỦA NGƯỜI DÙNG để tạo bảng Markdown từ nội dung phù hợp.\n"
            f"- Gợi ý: tối đa khoảng {int(self.valves.max_rows_hint)} dòng dữ liệu.\n\n"
            + "\n\n".join(lines)
        )

    # =========================================================
    # Markdown table parsing
    # =========================================================

    def _is_table_line(self, line: str) -> bool:
        return line.count("|") >= 2 and len(line.strip()) >= 3

    def _is_separator_line(self, line: str) -> bool:
        s = line.strip().strip("|").strip()
        if not s:
            return False
        return all(ch in "-: |" for ch in line.strip()) and "-" in line

    def _split_row(self, line: str) -> List[str]:
        return [c.strip() for c in line.strip().strip("|").split("|")]

    def _extract_tables(self, text: str) -> List[List[str]]:
        lines = text.splitlines()
        tables: List[List[str]] = []
        i = 0
        n = len(lines)
        while i < n:
            if (
                self._is_table_line(lines[i])
                and i + 1 < n
                and self._is_separator_line(lines[i + 1])
            ):
                block = [lines[i].strip(), lines[i + 1].strip()]
                i += 2
                while i < n and self._is_table_line(lines[i]):
                    block.append(lines[i].strip())
                    i += 1
                tables.append(block)
                continue
            i += 1
        return tables

    def _table_block_to_rows(self, block_lines: List[str]) -> List[List[str]]:
        if len(block_lines) < 2:
            return []
        header = self._split_row(block_lines[0])
        rows = [header]
        for ln in block_lines[2:]:
            rows.append(self._split_row(ln))

        max_cols = max(len(r) for r in rows) if rows else 0
        for r in rows:
            if len(r) < max_cols:
                r += [""] * (max_cols - len(r))
        return rows

    # =========================================================
    # Type inference
    # =========================================================

    _re_int = re.compile(r"^-?\d+$")
    _re_float = re.compile(r"^-?\d+(\.\d+)?$")
    _re_thousand = re.compile(r"^-?\d{1,3}([.,]\d{3})+(\.\d+)?$")
    _re_percent = re.compile(r"^-?\d+(\.\d+)?%$")
    _re_money = re.compile(r"^[₫$€]?\s*-?\d{1,3}([.,]\d{3})+(\.\d+)?\s*[₫$€]?$")
    _re_date1 = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    _re_date2 = re.compile(r"^\d{2}/\d{2}/\d{4}$")

    def _parse_number(self, s: str) -> Optional[float]:
        t = s.strip().replace(" ", "")
        if self._re_percent.match(t):
            try:
                return float(t[:-1]) / 100.0
            except Exception:
                return None

        t2 = t.replace("₫", "").replace("$", "").replace("€", "")
        if self._re_thousand.match(t2) or self._re_money.match(t):
            if "," in t2 and "." in t2:
                t2 = t2.replace(",", "")
            else:
                t2 = t2.replace(".", "").replace(",", "")
            try:
                return float(t2)
            except Exception:
                return None

        if self._re_int.match(t2):
            try:
                return int(t2)
            except Exception:
                return None

        if self._re_float.match(t2):
            try:
                return float(t2)
            except Exception:
                return None

        return None

    def _maybe_parse_date(self, s: str):
        t = s.strip()
        try:
            import datetime as _dt

            if self._re_date1.match(t):
                y, m, d = t.split("-")
                return _dt.date(int(y), int(m), int(d))
            if self._re_date2.match(t):
                d, m, y = t.split("/")
                return _dt.date(int(y), int(m), int(d))
        except Exception:
            return None
        return None

    # =========================================================
    # Excel builder
    # =========================================================

    def _autosize_columns(self, ws, min_w=10, max_w=55):
        max_w = int(self.valves.max_col_width) if self.valves.max_col_width else max_w
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            width = min(max(min_w, max_len + 2), max_w)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _apply_borders(self, ws, start_row, start_col, end_row, end_col):
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(r, c).border = border

    def _build_xlsx_bytes(self, markdown_text: str) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = self.valves.sheet_name

        ws["A1"] = self.valves.title
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = (
            f"Thời gian tạo: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        ws["A2"].font = Font(size=10, color="666666")

        tables = self._extract_tables(markdown_text)
        chosen = (
            tables[self.valves.table_index]
            if len(tables) > self.valves.table_index
            else None
        )
        start_row = int(self.valves.table_start_row)

        if not chosen:
            ws[f"A{start_row}"] = (
                "Không tìm thấy Markdown table trong nội dung để xuất."
            )
            ws[f"A{start_row}"].font = Font(bold=True, color="C00000")
            ws[f"A{start_row+2}"] = "raw_text"
            ws[f"A{start_row+2}"].font = Font(bold=True)
            ws[f"A{start_row+3}"] = markdown_text
            ws[f"A{start_row+3}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions["A"].width = 90

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio.getvalue()

        rows = self._table_block_to_rows(chosen)
        if not rows or len(rows) < 1:
            raise ValueError("Bảng Markdown không hợp lệ (không có header).")

        headers = rows[0]
        data_rows = rows[1:]

        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=c_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        for r_off, row in enumerate(data_rows, start=1):
            for c_idx, v in enumerate(row, start=1):
                val = v.strip() if isinstance(v, str) else (v if v is not None else "")

                d = self._maybe_parse_date(val) if isinstance(val, str) else None
                if d is not None:
                    cell = ws.cell(row=start_row + r_off, column=c_idx, value=d)
                    cell.number_format = "yyyy-mm-dd"
                else:
                    num = self._parse_number(val) if isinstance(val, str) else None
                    if num is not None:
                        cell = ws.cell(row=start_row + r_off, column=c_idx, value=num)
                        if isinstance(val, str) and val.strip().endswith("%"):
                            cell.number_format = "0.00%"
                        else:
                            cell.number_format = (
                                "#,##0.00"
                                if (
                                    isinstance(num, float)
                                    and not float(num).is_integer()
                                )
                                else "#,##0"
                            )
                    else:
                        cell = ws.cell(row=start_row + r_off, column=c_idx, value=val)

                cell.alignment = Alignment(vertical="top", wrap_text=True)

        end_row = start_row + len(data_rows)
        end_col = len(headers)

        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        self._apply_borders(ws, start_row, 1, end_row, end_col)

        table_ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
        tname = f"T_{datetime.datetime.now().strftime('%H%M%S')}"
        tab = Table(displayName=tname, ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

        self._autosize_columns(ws)
        ws.row_dimensions[start_row].height = 22

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.getvalue()

    # =========================================================
    # Open WebUI call (forward Authorization or Cookie)
    # =========================================================

    def _resolve_openwebui_base_url(self, __request__=None) -> Optional[str]:
        if (
            isinstance(self.valves.openwebui_base_url, str)
            and self.valves.openwebui_base_url.strip()
        ):
            return self.valves.openwebui_base_url.strip().rstrip("/")

        if __request__ is not None:
            try:
                base = str(__request__.base_url).rstrip("/")
                if base:
                    return base
            except Exception:
                pass

            try:
                origin = __request__.headers.get("origin")
                if origin:
                    return origin.rstrip("/")
                host = __request__.headers.get("host")
                proto = __request__.headers.get("x-forwarded-proto", "http")
                if host:
                    return f"{proto}://{host}".rstrip("/")
            except Exception:
                pass

        return None

    def _extract_auth_headers_from_request(self, __request__=None) -> Dict[str, str]:
        headers: Dict[str, str] = {"Content-Type": "application/json"}
        if __request__ is None:
            return headers

        try:
            auth = __request__.headers.get("authorization")
            if auth:
                headers["Authorization"] = auth
        except Exception:
            pass

        try:
            cookie = __request__.headers.get("cookie")
            if cookie and "Authorization" not in headers:
                headers["Cookie"] = cookie
        except Exception:
            pass

        return headers

    async def _requests_post_json(
        self, url: str, headers: Dict[str, str], payload: Dict[str, Any], timeout: int
    ):
        def _do():
            return requests.post(url, headers=headers, json=payload, timeout=timeout)

        return await asyncio.to_thread(_do)

    async def _normalize_markdown_via_openwebui(
        self,
        body: dict,
        __model__=None,
        __request__=None,
    ) -> str:
        base = self._resolve_openwebui_base_url(__request__)
        if not base:
            raise ValueError(
                "Không xác định được Open WebUI base_url. Hãy set valves.openwebui_base_url (VD http://localhost:3000)."
            )

        url = base + self.valves.openwebui_chat_completions_path
        headers = self._extract_auth_headers_from_request(__request__)

        model_id = self._get_current_model_id(body, __model__)
        if not model_id:
            raise ValueError(
                "Không xác định được model đang dùng trong chat để gọi normalize."
            )

        user_prompt = self._build_normalize_user_prompt(body)

        messages = [
            {"role": "system", "content": self.valves.normalize_instructions_vi},
            {"role": "user", "content": user_prompt},
        ]

        payload = {"model": model_id, "messages": messages, "stream": False}
        resp = await self._requests_post_json(
            url, headers=headers, payload=payload, timeout=int(self.valves.timeout_sec)
        )

        if resp.status_code >= 400:
            raise ValueError(
                f"Normalize failed: HTTP {resp.status_code} - {resp.text[:500]}"
            )

        data = resp.json()
        content = None
        try:
            content = data.get("choices", [{}])[0].get("message", {}).get("content")
        except Exception:
            content = None

        if not content or not isinstance(content, str) or not content.strip():
            content = (
                data.get("message", {}).get("content") or data.get("content") or ""
            )

        if not content.strip():
            raise ValueError(
                "Normalize: không lấy được content từ /api/chat/completions response."
            )

        return content

    # =========================================================
    # Main action
    # =========================================================

    async def action(
        self,
        body: dict,
        __user__=None,
        __event_emitter__=None,
        __event_call__=None,
        __model__=None,
        __request__=None,
        __id__=None,
    ):
        wizard_start = asyncio.get_event_loop().time()

        scope_text = (
            "assistant_only"
            if (self.valves.normalize_mode or "").strip() == "assistant_only"
            else "conversation_window"
        )
        steps = [
            f"1) Chuẩn hoá bảng theo mode: {scope_text}",
            "2) Tạo file Excel (định dạng, filter, freeze header)",
            "3) Tải file xuống trình duyệt",
        ]

        try:
            if self.valves.show_confirmation and __event_call__:
                ok = await __event_call__(
                    {
                        "type": "confirmation",
                        "data": {
                            "title": "Xuất Excel",
                            "message": "Bắt đầu chuẩn hoá & xuất Excel từ hội thoại hiện tại?",
                        },
                    }
                )
                if not ok:
                    if __event_emitter__:
                        await __event_emitter__(
                            {
                                "type": "notification",
                                "data": {
                                    "type": "info",
                                    "content": "Đã huỷ thao tác xuất Excel.",
                                },
                            }
                        )
                    return body

            await self._ui_modal_open(__event_call__)
            await self._ui_modal_update(
                __event_call__, subtitle="Đang khởi tạo...", steps=steps
            )

            if __event_emitter__ and self.valves.show_status:
                await __event_emitter__(
                    {
                        "type": "notification",
                        "data": {"type": "info", "content": "Đang xuất Excel..."},
                    }
                )
                await __event_emitter__(
                    {
                        "type": "status",
                        "data": {"description": "Bắt đầu xử lý...", "done": False},
                    }
                )

            # Step 1: Normalize
            await self._ui_modal_update(
                __event_call__,
                subtitle="Bước 1/3: Đang chuẩn hoá nội dung thành bảng Markdown...",
                steps=steps,
            )
            if self.valves.normalize_with_llm:
                markdown_text = await self._normalize_markdown_via_openwebui(
                    body=body, __model__=__model__, __request__=__request__
                )
            else:
                raise ValueError(
                    "normalize_with_llm đang tắt. Hãy bật để tự chuẩn hoá bảng theo context."
                )

            # Step 2: Build Excel
            await self._ui_modal_update(
                __event_call__, subtitle="Bước 2/3: Đang tạo file Excel...", steps=steps
            )
            if __event_emitter__ and self.valves.show_status:
                await __event_emitter__(
                    {
                        "type": "status",
                        "data": {
                            "description": "Đang tạo file Excel...",
                            "done": False,
                        },
                    }
                )

            xlsx_bytes = self._build_xlsx_bytes(markdown_text)

            # Step 3: Download
            await self._ui_modal_update(
                __event_call__, subtitle="Bước 3/3: Đang tải file xuống...", steps=steps
            )

            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{ts}_{self.valves.base_filename}"
            xlsx_b64 = base64.b64encode(xlsx_bytes).decode("utf-8")

            js = f"""
try {{
  const base64Data = "{xlsx_b64}";
  const binaryData = atob(base64Data);
  const bytes = new Uint8Array(binaryData.length);
  for (let i = 0; i < binaryData.length; i++) bytes[i] = binaryData.charCodeAt(i);
  const blob = new Blob([bytes], {{ type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" }});
  const url = URL.createObjectURL(blob);
  const a = document.createElement("a");
  a.href = url;
  a.download = "{filename}";
  document.body.appendChild(a);
  a.click();
  URL.revokeObjectURL(url);
  a.remove();
}} catch (e) {{
  console.error("Download XLSX failed", e);
}}
"""
            if __event_call__:
                await __event_call__({"type": "execute", "data": {"code": js}})

            # Ensure wizard visible
            elapsed = asyncio.get_event_loop().time() - wizard_start
            min_s = float(self.valves.min_wizard_seconds or 0.0)
            if elapsed < min_s:
                await asyncio.sleep(min_s - elapsed)

            await self._ui_modal_update(__event_call__, done_ok=True, steps=steps)

            if __event_emitter__ and self.valves.show_status:
                await __event_emitter__(
                    {
                        "type": "status",
                        "data": {
                            "description": "Hoàn tất. File đang được tải xuống.",
                            "done": True,
                        },
                    }
                )
                await __event_emitter__(
                    {
                        "type": "notification",
                        "data": {
                            "type": "success",
                            "content": "Đã xuất Excel thành công.",
                        },
                    }
                )

            return body

        except Exception as e:
            logger.exception("Lỗi xuất Excel: %s", str(e))

            elapsed = asyncio.get_event_loop().time() - wizard_start
            min_s = float(self.valves.min_wizard_seconds or 0.0)
            if elapsed < min_s:
                await asyncio.sleep(min_s - elapsed)

            await self._ui_modal_update(
                __event_call__, done_ok=False, error_text=str(e), steps=steps
            )

            if __event_emitter__:
                await __event_emitter__(
                    {
                        "type": "status",
                        "data": {"description": f"Lỗi xuất Excel: {e}", "done": True},
                    }
                )
                await __event_emitter__(
                    {
                        "type": "notification",
                        "data": {"type": "error", "content": f"Lỗi xuất Excel: {e}"},
                    }
                )

            return None
