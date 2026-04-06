"""
Script tạo Form Test Excel cho Open WebUI
Chạy: python generate_test_excel.py
Output: FORM_TEST_OPEN_WEBUI.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ── Styles ──────────────────────────────────────────────────────────
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name='Arial', size=16, bold=True, color='FFFFFF')
TITLE_FILL = PatternFill('solid', fgColor='1F4E79')

SECTION_FONT = Font(name='Arial', size=13, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill('solid', fgColor='2E75B6')

SUBSECTION_FONT = Font(name='Arial', size=11, bold=True, color='1F4E79')
SUBSECTION_FILL = PatternFill('solid', fgColor='D6E4F0')

HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')

DATA_FONT = Font(name='Arial', size=10)
DATA_FILL_ODD = PatternFill('solid', fgColor='F2F7FB')
DATA_FILL_EVEN = PatternFill('solid', fgColor='FFFFFF')

PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
PARTIAL_FILL = PatternFill('solid', fgColor='FFEB9C')

SUMMARY_FONT = Font(name='Arial', size=11, bold=True)
SUMMARY_FILL = PatternFill('solid', fgColor='E2EFDA')

WRAP = Alignment(wrap_text=True, vertical='center')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

COLS = ['STT', 'Nhóm tính năng', 'Câu hỏi / Thao tác test', 'Kỳ vọng',
        'Câu trả lời thực tế', 'Kết quả', 'Đánh giá (1-5)', 'Ghi chú / Bug']
COL_WIDTHS = [6, 22, 45, 40, 40, 12, 14, 30]

# ── Test Data ───────────────────────────────────────────────────────
SECTIONS = [
    {
        'title': 'PHẦN A: TÍNH NĂNG NGƯỜI DÙNG (USER)',
        'groups': [
            {
                'name': 'A1. HỎI ĐÁP (CHAT AI)',
                'subs': [
                    {
                        'sub': 'A1.1 Chất lượng trả lời',
                        'rows': [
                            ['Trả lời tiếng Việt', 'Hỏi: "Giải thích nguyên lý hoạt động của pin mặt trời"', 'Trả lời mượt mà bằng tiếng Việt, nội dung chính xác, không lỗi ngữ pháp'],
                            ['Trả lời tiếng Anh', 'Hỏi: "Explain how LED lighting works"', 'Trả lời tiếng Anh chính xác, ngữ pháp tốt'],
                            ['Trả lời chuyên ngành', 'Hỏi: "So sánh ưu nhược điểm đèn LED và đèn huỳnh quang trong chiếu sáng công nghiệp"', 'Thông tin chuyên ngành chính xác, có so sánh rõ'],
                            ['Trả lời có số liệu', 'Hỏi: "Dân số Việt Nam năm 2025 là bao nhiêu?"', 'Số liệu gần đúng, có disclaimer'],
                            ['Trả lời dạng bảng', 'Hỏi: "Lập bảng so sánh 5 loại đèn LED phổ biến"', 'Hiển thị bảng markdown đẹp, dữ liệu hợp lý'],
                            ['Trả lời dạng code', 'Hỏi: "Viết code Python đọc file Excel và vẽ biểu đồ"', 'Code chạy được, có syntax highlighting'],
                            ['Hội thoại nhiều lượt', 'Hỏi 3-4 câu liên tiếp về cùng 1 chủ đề', 'AI nhớ context, trả lời liên tục mạch lạc'],
                            ['Streaming response', 'Quan sát khi AI trả lời', 'Text hiển thị từng token, không bị giật lag'],
                        ]
                    },
                    {
                        'sub': 'A1.2 Chức năng hội thoại',
                        'rows': [
                            ['Tạo hội thoại mới', 'Nhấn nút "New Chat"', 'Mở chat mới, không còn context cũ'],
                            ['Xóa hội thoại', 'Xóa 1 hội thoại từ sidebar', 'Hội thoại bị xóa khỏi danh sách'],
                            ['Xóa nhiều hội thoại', 'Xóa 3 hội thoại cùng lúc', 'Tất cả bị xóa thành công'],
                            ['Đổi tên hội thoại', 'Click tên → Sửa tên', 'Tên mới được lưu, hiển thị đúng'],
                            ['Pin hội thoại', 'Ghim 1 hội thoại', 'Hiển thị ở đầu danh sách'],
                            ['Archive hội thoại', 'Lưu trữ 1 hội thoại', 'Chuyển vào mục Archive'],
                            ['Tạo folder', 'Tạo folder mới trong sidebar', 'Folder được tạo, kéo thả chat vào được'],
                            ['Tìm kiếm chat', 'Gõ keyword vào thanh search', 'Tìm được hội thoại chứa keyword'],
                            ['Chia sẻ hội thoại', 'Nhấn Share → Copy link', 'Link được tạo, người khác mở được'],
                            ['Xuất file Excel', 'Action → "Xuất Excel"', 'File .xlsx tải về, có format, filter'],
                            ['Xuất file PDF', 'Action → "Xuất PDF"', 'File .pdf tải về, nội dung đúng'],
                            ['Xuất file DOCX', 'Action → "Xuất DOCX"', 'File .docx tải về, nội dung đúng'],
                            ['Tags', 'Gắn tag cho hội thoại', 'Tag được lưu, lọc theo tag hoạt động'],
                        ]
                    },
                ]
            },
            {
                'name': 'A2. TẠO ẢNH AI',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Tạo ảnh DALL-E 3', 'Chọn img-gpt-dalle-3, gõ mô tả ảnh', 'Ảnh được tạo, hiển thị đúng mô tả'],
                            ['Tạo ảnh Gemini Flash', 'Chọn img-gemini-flash, gõ mô tả', 'Ảnh được tạo nhanh, chất lượng OK'],
                            ['Tạo ảnh Gemini Pro', 'Chọn img-gemini-pro, gõ mô tả', 'Ảnh chất lượng cao, chi tiết tốt'],
                            ['Sửa ảnh qua nhiều lượt', 'Lượt 1: Vẽ → Lượt 2: Thêm chi tiết → Lượt 3: Đổi style', 'Mỗi lượt chỉnh sửa dựa trên ảnh trước'],
                            ['Tạo ảnh phức tạp', 'Gõ mô tả dài, nhiều chi tiết', 'Ảnh phản ánh đầy đủ mô tả'],
                            ['Tạo ảnh prompt tiếng Việt', 'Gõ prompt bằng tiếng Việt', 'Hiểu và tạo ảnh đúng mô tả TV'],
                        ]
                    }
                ]
            },
            {
                'name': 'A3. GỬI FILE & TRUY XUẤT THÔNG TIN',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Upload PDF', 'Gửi 1 file PDF (~5 trang) + hỏi nội dung', 'AI đọc và trả lời chính xác'],
                            ['Upload Word', 'Gửi 1 file .docx + hỏi tóm tắt', 'Tóm tắt đúng, đầy đủ ý chính'],
                            ['Upload Excel', 'Gửi file Excel có số liệu + hỏi tổng', 'Đọc được số liệu, trả lời đúng'],
                            ['Upload CSV', 'Gửi 1 file CSV + hỏi thông tin', 'Đọc đúng, trả lời chính xác'],
                            ['Upload TXT', 'Gửi 1 file .txt + hỏi nội dung', 'Đọc và trả lời đúng'],
                            ['Upload nhiều file', 'Gửi 3 file khác định dạng cùng lúc', 'Xử lý được tất cả'],
                            ['Upload file lớn', 'Gửi file PDF ~20-30MB', 'Upload thành công hoặc báo lỗi rõ'],
                            ['Hỏi đáp sâu file', 'Upload → hỏi 3-4 câu liên tiếp', 'AI nhớ file, trả lời chính xác'],
                            ['Upload HTML', 'Gửi 1 file .html + hỏi', 'Đọc được nội dung text từ HTML'],
                            ['Trích dẫn nguồn', 'Hỏi thông tin từ file đã upload', 'AI trích dẫn tên file, trang nguồn'],
                        ]
                    }
                ]
            },
            {
                'name': 'A4. CHỌN MODEL',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Danh sách model', 'Mở dropdown chọn model', 'Hiển thị đầy đủ tất cả model'],
                            ['GPT-5', 'Chọn GPT-5 → hỏi 1 câu', 'Trả lời thành công, chất lượng cao'],
                            ['GPT-5 Mini', 'Chọn GPT-5 Mini → hỏi 1 câu', 'Trả lời nhanh, chất lượng tốt'],
                            ['GPT-4o', 'Chọn GPT-4o → hỏi 1 câu', 'Trả lời thành công'],
                            ['GPT-4.1', 'Chọn GPT-4.1 → hỏi 1 câu', 'Trả lời thành công'],
                            ['Gemini 2.5 Pro', 'Chọn Gemini 2.5 Pro → hỏi 1 câu', 'Trả lời thành công, TV tốt'],
                            ['Gemini 2.5 Flash', 'Chọn Gemini 2.5 Flash → hỏi 1 câu', 'Trả lời nhanh, chất lượng ổn'],
                            ['Gemini 3 Pro', 'Chọn Gemini 3 Pro → hỏi 1 câu', 'Flagship mới nhất, trả lời tốt'],
                            ['Chuyển model giữa chừng', 'Chat GPT-5 → đổi Gemini → hỏi tiếp', 'Hội thoại tiếp tục đúng'],
                            ['Model mặc định', 'Settings → chọn model mặc định', 'Chat mới tự động dùng model đã chọn'],
                        ]
                    }
                ]
            },
            {
                'name': 'A5. KHO KIẾN THỨC (KNOWLEDGE)',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Tạo Knowledge', 'Workspace → Knowledge → Create', 'Collection được tạo thành công'],
                            ['Upload PDF vào Knowledge', 'Upload file PDF (~5MB)', 'Upload thành công, được index'],
                            ['Upload Word vào Knowledge', 'Upload file .docx', 'Upload thành công'],
                            ['Upload Excel vào Knowledge', 'Upload file .xlsx', 'Upload thành công'],
                            ['Upload CSV vào Knowledge', 'Upload file .csv', 'Upload thành công'],
                            ['Upload nhiều file', 'Upload 5+ file khác nhau', 'Tất cả upload thành công'],
                            ['Upload file lớn', 'Upload file ~30-50MB', 'Thành công hoặc báo lỗi rõ ràng'],
                            ['Gọi Knowledge #', 'Trong chat gõ #tên-knowledge', 'Hiển thị gợi ý, chọn được'],
                            ['Hỏi đáp từ Knowledge', 'Gọi # rồi hỏi nội dung', 'Trả lời dựa trên Knowledge, có trích dẫn'],
                            ['Xóa file', 'Xóa 1 file trong Knowledge', 'File + embedding bị xóa'],
                            ['Xóa Knowledge', 'Xóa cả collection', 'Collection + tất cả bị xóa'],
                        ]
                    }
                ]
            },
            {
                'name': 'A6. WORKSPACE',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Tạo Workspace', 'Tạo Workspace mới', 'Workspace được tạo thành công'],
                            ['Chọn Workspace', 'Chuyển đổi giữa các Workspace', 'Chuyển đổi mượt, dữ liệu riêng biệt'],
                            ['Chia sẻ Knowledge', 'Chia sẻ Knowledge cho user khác', 'User khác truy cập được'],
                            ['Access control', 'Giới hạn quyền truy cập', 'Chỉ user được phép mới truy cập'],
                        ]
                    }
                ]
            },
            {
                'name': 'A7. GIỌNG NÓI (VOICE)',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Text-to-Speech', 'Nhấn icon 🔊 trên response', 'AI đọc bằng giọng nói tự nhiên'],
                            ['Speech-to-Text', 'Nhấn icon 🎤 → nói câu hỏi', 'Giọng nói chuyển thành text chính xác'],
                            ['STT tiếng Việt', 'Nói bằng tiếng Việt', 'Nhận dạng tiếng Việt chính xác'],
                        ]
                    }
                ]
            },
            {
                'name': 'A8. TÌM KIẾM WEB',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Web search cơ bản', 'Hỏi: "Tin tức mới nhất về AI"', 'Search web, trả lời có nguồn trích dẫn'],
                            ['Trích dẫn nguồn web', 'Kiểm tra response có tag nguồn', 'Hiển thị chip/tag website + URL'],
                            ['Search tiếng Việt', 'Hỏi: "Thời tiết Hà Nội hôm nay"', 'Tìm và trả lời TV, có nguồn'],
                        ]
                    }
                ]
            },
            {
                'name': 'A9. CÁ NHÂN HÓA & GIAO DIỆN',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Dark/Light mode', 'Settings → Theme → chuyển đổi', 'Giao diện chuyển đổi mượt'],
                            ['Ngôn ngữ', 'Settings → Language → Tiếng Việt', 'Toàn bộ UI chuyển tiếng Việt'],
                            ['System prompt', 'Settings → đặt system prompt', 'AI tuân theo khi chat'],
                            ['Responsive mobile', 'Mở trên điện thoại', 'Hiển thị tốt, không vỡ layout'],
                        ]
                    }
                ]
            },
        ]
    },
    {
        'title': 'PHẦN B: TÍNH NĂNG QUẢN TRỊ (ADMIN)',
        'groups': [
            {
                'name': 'B1. QUẢN LÝ NGƯỜI DÙNG',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Xem danh sách user', 'Admin Panel → Users', 'Hiển thị đầy đủ danh sách, role, trạng thái'],
                            ['Tạo user mới', 'Tạo account / đăng ký', 'Account tạo, trạng thái Pending'],
                            ['Duyệt user', 'Approve user pending', 'User Active, đăng nhập được'],
                            ['Đổi role', 'User → Admin', 'Role + quyền thay đổi'],
                            ['Xóa user', 'Xóa 1 user account', 'User bị xóa, không login được'],
                            ['Reset password', 'Reset password cho user', 'Password mới hoạt động'],
                        ]
                    }
                ]
            },
            {
                'name': 'B2. QUẢN LÝ MODEL',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Bật/tắt model', 'Admin → Models → Disable 1 model', 'User không thấy model đã disable'],
                            ['Cấu hình params', 'Set temperature, max_tokens', 'Params áp dụng khi user chat'],
                            ['Gán Knowledge model', 'Gán Knowledge mặc định cho model', 'Model tự động dùng Knowledge'],
                            ['Access control model', 'Giới hạn model cho user cụ thể', 'Chỉ user được phép mới thấy'],
                        ]
                    }
                ]
            },
            {
                'name': 'B3. DASHBOARD & GIÁM SÁT',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Dashboard tổng quan', 'Truy cập Dashboard middleware', 'Hiển thị tổng chi phí, request, top users'],
                            ['Log requests', 'Xem log chi tiết API request', 'Hiển thị user, model, tokens, cost'],
                            ['Báo cáo theo user', 'Lọc chi phí theo user', 'Dữ liệu chính xác'],
                            ['Báo cáo theo model', 'Lọc chi phí theo model', 'Dữ liệu chính xác'],
                            ['Báo cáo theo thời gian', 'Lọc theo khoảng thời gian', 'Dữ liệu đúng trong khoảng chọn'],
                        ]
                    }
                ]
            },
            {
                'name': 'B4. QUOTA & CHI PHÍ',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['Set quota user', 'Dashboard → set limit_cost_usd', 'Quota được lưu thành công'],
                            ['Cảnh báo quota', 'User chat gần hết quota', 'Hiển thị cảnh báo'],
                            ['Chặn khi hết quota', 'User vượt quota → chat tiếp', 'Từ chối, báo lỗi HTTP 429'],
                            ['Sub-key API', 'Kiểm tra user có sub-key riêng', 'Mỗi user có API key riêng'],
                        ]
                    }
                ]
            },
            {
                'name': 'B5. CẤU HÌNH HỆ THỐNG',
                'subs': [
                    {
                        'sub': None,
                        'rows': [
                            ['WebUI settings', 'Admin → Settings → General', 'Cấu hình được lưu và áp dụng'],
                            ['Connections', 'Admin → Settings → Connections', 'Hiển thị kết nối providers, status OK'],
                            ['RAG config', 'Điều chỉnh chunk size, file limit', 'Thay đổi được áp dụng'],
                            ['Signup control', 'Bật/tắt đăng ký mới', 'Phản ánh đúng trạng thái bật/tắt'],
                        ]
                    }
                ]
            },
        ]
    }
]


def apply_style(cell, font=None, fill=None, alignment=None, border=None):
    if font:     cell.font = font
    if fill:     cell.fill = fill
    if alignment: cell.alignment = alignment
    if border:   cell.border = border


def build_sheet(ws):
    ws.sheet_properties.tabColor = '1F4E79'
    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    # ── Title ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value='FORM KIỂM THỬ TÍNH NĂNG - OPEN WEBUI')
    apply_style(c, TITLE_FONT, TITLE_FILL, CENTER)
    for col in range(2, 9):
        apply_style(ws.cell(row=row, column=col), fill=TITLE_FILL)
    ws.row_dimensions[row].height = 40
    row += 1

    # ── Info row ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    info = 'Phiên bản: v1.0  |  Ngày: 2026-04-01  |  URL: https://openwebui.rangdong.com.vn'
    c = ws.cell(row=row, column=1, value=info)
    apply_style(c, Font(name='Arial', size=10, italic=True, color='666666'), alignment=CENTER)
    row += 2

    stt = 1  # Global STT counter
    summary_data = []  # For summary sheet

    for section in SECTIONS:
        # ── Section Title (PHẦN A / B) ──
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=section['title'])
        apply_style(c, SECTION_FONT, SECTION_FILL, Alignment(horizontal='left', vertical='center'))
        for col in range(2, 9):
            apply_style(ws.cell(row=row, column=col), fill=SECTION_FILL)
        ws.row_dimensions[row].height = 32
        row += 1

        for group in section['groups']:
            group_count = sum(len(s['rows']) for s in group['subs'])
            group_start_stt = stt

            # ── Group title (A1, A2...) ──
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value=group['name'])
            apply_style(c, SUBSECTION_FONT, SUBSECTION_FILL,
                        Alignment(horizontal='left', vertical='center'))
            for col in range(2, 9):
                apply_style(ws.cell(row=row, column=col), fill=SUBSECTION_FILL)
            ws.row_dimensions[row].height = 26
            row += 1

            for sub_group in group['subs']:
                # ── Sub-group title (A1.1, A1.2...) ──
                if sub_group.get('sub'):
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                    c = ws.cell(row=row, column=1, value=f"  ▸ {sub_group['sub']}")
                    apply_style(c, Font(name='Arial', size=10, bold=True, italic=True, color='2E75B6'),
                                PatternFill('solid', fgColor='EBF1F8'),
                                Alignment(horizontal='left', vertical='center'))
                    for col in range(2, 9):
                        apply_style(ws.cell(row=row, column=col),
                                    fill=PatternFill('solid', fgColor='EBF1F8'))
                    row += 1

                # ── Header row ──
                for i, h in enumerate(COLS, 1):
                    c = ws.cell(row=row, column=i, value=h)
                    apply_style(c, HEADER_FONT, HEADER_FILL, CENTER, BORDER)
                ws.row_dimensions[row].height = 24
                row += 1

                # ── Data rows ──
                for idx, r in enumerate(sub_group['rows']):
                    fill = DATA_FILL_ODD if idx % 2 == 0 else DATA_FILL_EVEN
                    vals = [stt, r[0], r[1], r[2], '', '', '', '']
                    for i, v in enumerate(vals, 1):
                        c = ws.cell(row=row, column=i, value=v)
                        al = CENTER if i in (1, 6, 7) else WRAP
                        apply_style(c, DATA_FONT, fill, al, BORDER)
                    ws.row_dimensions[row].height = 45
                    stt += 1
                    row += 1

            summary_data.append((group['name'], group_count))
            row += 1  # gap between groups

    return summary_data, row


def build_summary(ws, summary_data):
    ws.sheet_properties.tabColor = '548235'
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 14

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value='BẢNG TỔNG HỢP KẾT QUẢ TEST')
    apply_style(c, TITLE_FONT, TITLE_FILL, CENTER)
    for col in range(2, 9):
        apply_style(ws.cell(row=row, column=col), fill=TITLE_FILL)
    ws.row_dimensions[row].height = 38
    row += 2

    headers = ['STT', 'Nhóm tính năng', 'Tổng test case', 'PASS', 'FAIL', 'PARTIAL', 'SKIP', 'Tỷ lệ PASS']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        apply_style(c, HEADER_FONT, HEADER_FILL, CENTER, BORDER)
    row += 1

    total = 0
    for idx, (name, count) in enumerate(summary_data, 1):
        fill = DATA_FILL_ODD if idx % 2 == 1 else DATA_FILL_EVEN
        vals = [idx, name, count, '', '', '', '', '']
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=v)
            al = CENTER if i != 2 else WRAP
            apply_style(c, DATA_FONT, fill, al, BORDER)
        total += count
        row += 1

    # Total row
    vals = ['', 'TỔNG CỘNG', total, '', '', '', '', '']
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        apply_style(c, SUMMARY_FONT, SUMMARY_FILL, CENTER, BORDER)
    row += 2

    # Tester info
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value='THÔNG TIN NGƯỜI TEST')
    apply_style(c, SUBSECTION_FONT, SUBSECTION_FILL, Alignment(horizontal='left', vertical='center'))
    for col in range(2, 9):
        apply_style(ws.cell(row=row, column=col), fill=SUBSECTION_FILL)
    row += 1

    info_fields = ['Họ và tên', 'Phòng ban', 'Ngày test', 'Trình duyệt sử dụng',
                   'Thiết bị (PC/Mobile)', 'Nhận xét chung', 'Đề xuất cải thiện']
    for field in info_fields:
        c1 = ws.cell(row=row, column=1, value=field)
        apply_style(c1, Font(name='Arial', size=10, bold=True), border=BORDER, alignment=WRAP)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        c2 = ws.cell(row=row, column=3, value='')
        apply_style(c2, border=BORDER)
        for col in range(4, 9):
            apply_style(ws.cell(row=row, column=col), border=BORDER)
        ws.row_dimensions[row].height = 28
        row += 1


def build_guide(ws):
    ws.sheet_properties.tabColor = 'FFC000'
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 55

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='HƯỚNG DẪN SỬ DỤNG FORM TEST')
    apply_style(c, TITLE_FONT, TITLE_FILL, CENTER)
    for col in range(2, 4):
        apply_style(ws.cell(row=row, column=col), fill=TITLE_FILL)
    ws.row_dimensions[row].height = 38
    row += 2

    guides = [
        ['Ký hiệu', 'Ý nghĩa'],
        ['✅ PASS', 'Tính năng hoạt động đúng kỳ vọng'],
        ['❌ FAIL', 'Tính năng không hoạt động hoặc sai kỳ vọng'],
        ['⚠️ PARTIAL', 'Hoạt động một phần, cần ghi chú chi tiết'],
        ['⏳ SKIP', 'Chưa test / Bỏ qua'],
    ]
    for i, (k, v) in enumerate(guides):
        c1 = ws.cell(row=row, column=2, value=k)
        c2 = ws.cell(row=row, column=3, value=v)
        if i == 0:
            apply_style(c1, HEADER_FONT, HEADER_FILL, CENTER, BORDER)
            apply_style(c2, HEADER_FONT, HEADER_FILL, CENTER, BORDER)
        else:
            apply_style(c1, Font(name='Arial', size=11, bold=True), border=BORDER, alignment=CENTER)
            apply_style(c2, DATA_FONT, border=BORDER, alignment=WRAP)
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='CÁC CỘT CẦN ĐIỀN:')
    apply_style(c, SUBSECTION_FONT, SUBSECTION_FILL, Alignment(horizontal='left', vertical='center'))
    for col in range(2, 4):
        apply_style(ws.cell(row=row, column=col), fill=SUBSECTION_FILL)
    row += 1

    fill_cols = [
        ['Câu trả lời thực tế', 'Ghi lại response thực tế của AI hoặc kết quả thao tác'],
        ['Kết quả', 'Chọn: PASS / FAIL / PARTIAL / SKIP'],
        ['Đánh giá (1-5)', '1=Rất kém, 2=Kém, 3=Trung bình, 4=Tốt, 5=Xuất sắc'],
        ['Ghi chú / Bug', 'Mô tả bug, lỗi, hoặc nhận xét chi tiết'],
    ]
    for k, v in fill_cols:
        c1 = ws.cell(row=row, column=2, value=k)
        c2 = ws.cell(row=row, column=3, value=v)
        apply_style(c1, Font(name='Arial', size=10, bold=True), border=BORDER, alignment=CENTER)
        apply_style(c2, DATA_FONT, border=BORDER, alignment=WRAP)
        row += 1


def main():
    wb = openpyxl.Workbook()

    # Sheet 1: Guide
    ws_guide = wb.active
    ws_guide.title = 'Hướng dẫn'
    build_guide(ws_guide)

    # Sheet 2: Test form
    ws_test = wb.create_sheet('Form Test')
    summary_data, _ = build_sheet(ws_test)

    # Sheet 3: Summary
    ws_summary = wb.create_sheet('Tổng hợp')
    build_summary(ws_summary, summary_data)

    # Freeze pane for test sheet
    ws_test.freeze_panes = 'A5'

    out = 'FORM_TEST_OPEN_WEBUI.xlsx'
    wb.save(out)
    print(f'✅ Đã tạo file: {out}')
    print(f'   - Sheet "Hướng dẫn": cách sử dụng form')
    print(f'   - Sheet "Form Test": 95 test cases (User + Admin)')
    print(f'   - Sheet "Tổng hợp": bảng tổng hợp + thông tin người test')


if __name__ == '__main__':
    main()
