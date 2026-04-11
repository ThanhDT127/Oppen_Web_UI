"""
Form Test Excel v2 - Open WebUI
Chạy: python generate_test_excel_v2.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── Styles ──
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='center')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)

TITLE_FONT = Font(name='Arial', size=14, bold=True, color='FFFFFF')
TITLE_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', size=9, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
DATA_FONT = Font(name='Arial', size=9)
BOLD9 = Font(name='Arial', size=9, bold=True)
SECTION_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')

# Group color bands (alternating for groups)
GROUP_COLORS = [
    'F2F7FB', 'FFFFFF', 'EBF5EB', 'FFF8E7', 'F5F0FF',
    'FFF0F0', 'F0FAFF', 'FFFBF0', 'F0FFF4', 'FFF5F8',
    'F5F5F5', 'FFFAE6', 'F0F8FF', 'FFF0F5',
]
SECTION_FILLS = {
    'USER': PatternFill('solid', fgColor='2E75B6'),
    'ADMIN': PatternFill('solid', fgColor='C0504D'),
}

# Conditional formatting fills
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
PARTIAL_FILL = PatternFill('solid', fgColor='FFEB9C')
SKIP_FILL = PatternFill('solid', fgColor='D9D9D9')

# ── Column defs for Sheet 2 ──
COLS2 = [
    ('STT', 5),
    ('Phân hệ', 10),
    ('Nhóm tính năng', 22),
    ('Tính năng cụ thể', 22),
    ('Thao tác test', 42),
    ('Kỳ vọng', 35),
    ('Kết quả thực tế', 35),
    ('Ảnh chụp\nmàn hình', 16),
    ('Kết quả', 11),
    ('Đánh giá\n(1-5)', 10),
    ('Thời gian test\n(ngày/tháng/năm)', 16),
    ('Nhận xét,\nđánh giá', 30),
    ('Feedback /\nĐề xuất', 30),
]

# ── Column defs for Sheet 3 ──
COLS3 = [
    ('STT', 5),
    ('Người test', 16),
    ('Phòng ban /\nNghiệp vụ', 18),
    ('Kịch bản test', 30),
    ('Model\nsử dụng', 18),
    ('Câu hỏi / Prompt', 42),
    ('File đính kèm\n(nếu có)', 18),
    ('Kỳ vọng', 30),
    ('Kết quả thực tế', 35),
    ('Ảnh chụp\nmàn hình', 16),
    ('Kết quả', 11),
    ('Thời gian test\n(ngày/tháng/năm)', 16),
    ('Nhận xét,\nđánh giá', 30),
    ('Feedback /\nĐề xuất', 30),
]

# ── Test Data ──
# Format: (phan_he, nhom, tinh_nang, thao_tac, ky_vong)
# Models sourced from prices.json (2026-03-26):
#   OpenAI:    gpt-5.4, gpt-5.2, gpt-5, gpt-image-1.5, gpt-image-1
#   Gemini:    gemini-3.1-pro-preview, gemini-3.1-flash-lite-preview, gemini-2.5-flash
#              gemini-3.1-flash-image-preview, gemini-3-pro-image-preview
#   xAI/Grok:  grok-4.20-reasoning, grok-4-1-fast-reasoning, grok-4-1-fast-non-reasoning
#              grok-imagine-image, grok-imagine-image-pro
#   Anthropic: claude-opus-4-6, claude-sonnet-4-6, claude-haiku-4-5
TEST_CASES = [
    # ──────────────── A1. Hỏi đáp - Chất lượng ────────────────
    ('USER', 'Hỏi đáp (Chat AI)', 'Trả lời tiếng Việt', 'Hỏi: "Giải thích nguyên lý hoạt động của pin mặt trời"', 'Trả lời mượt mà tiếng Việt, chính xác, không lỗi ngữ pháp'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Trả lời tiếng Anh', 'Hỏi: "Explain how LED lighting works"', 'Trả lời tiếng Anh chính xác, ngữ pháp tốt'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Trả lời chuyên ngành', 'Hỏi về so sánh đèn LED vs đèn huỳnh quang', 'Thông tin chuyên ngành chính xác, so sánh rõ ràng'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Trả lời có số liệu', 'Hỏi: "Dân số Việt Nam năm 2025?"', 'Số liệu gần đúng, có disclaimer'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Trả lời dạng bảng', 'Hỏi: "Lập bảng so sánh 5 loại đèn LED"', 'Bảng markdown đẹp, dữ liệu hợp lý'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Trả lời dạng code', 'Hỏi: "Viết code Python đọc Excel vẽ biểu đồ"', 'Code chạy được, syntax highlighting'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Hội thoại nhiều lượt', 'Hỏi 3-4 câu liên tiếp cùng chủ đề', 'AI nhớ context, trả lời mạch lạc'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Streaming response', 'Quan sát khi AI trả lời', 'Text hiển thị từng token, không giật lag'),
    # ──────────────── A1. Hỏi đáp - Chức năng ────────────────
    ('USER', 'Hỏi đáp (Chat AI)', 'Tạo hội thoại mới', 'Nhấn "New Chat"', 'Mở chat mới, không context cũ'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Xóa hội thoại', 'Xóa 1 hội thoại từ sidebar', 'Hội thoại bị xóa khỏi danh sách'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Xóa nhiều hội thoại', 'Xóa 3 hội thoại cùng lúc', 'Tất cả bị xóa thành công'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Đổi tên hội thoại', 'Click tên → Sửa tên', 'Tên mới được lưu'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Pin hội thoại', 'Ghim 1 hội thoại', 'Hiển thị ở đầu danh sách'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Archive hội thoại', 'Lưu trữ 1 hội thoại', 'Chuyển vào mục Archive'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Tạo folder', 'Tạo folder mới trong sidebar', 'Folder tạo, kéo thả chat vào được'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Tìm kiếm chat', 'Gõ keyword vào thanh search', 'Tìm được hội thoại chứa keyword'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Chia sẻ hội thoại', 'Nhấn Share → Copy link', 'Link được tạo, người khác mở được'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Xuất file Excel', 'Action → "Xuất Excel"', 'File .xlsx tải về, format + filter tốt'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Xuất file PDF', 'Action → "Xuất PDF"', 'File .pdf tải về, nội dung đúng'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Xuất file DOCX', 'Action → "Xuất DOCX"', 'File .docx tải về, nội dung đúng'),
    ('USER', 'Hỏi đáp (Chat AI)', 'Tags', 'Gắn tag cho hội thoại', 'Tag lưu, lọc theo tag hoạt động'),
    # ──────────────── A2. Tạo ảnh AI ────────────────
    ('USER', 'Tạo ảnh AI', 'GPT Image 1.5 (OpenAI)', 'Chọn model gpt-image-1.5, gõ: "Vẽ con mèo trên bàn làm việc"', 'Ảnh tạo đúng mô tả, chất lượng cao'),
    ('USER', 'Tạo ảnh AI', 'GPT Image 1 (OpenAI)', 'Chọn model gpt-image-1, gõ mô tả ảnh', 'Ảnh tạo đúng mô tả'),
    ('USER', 'Tạo ảnh AI', 'Gemini 3.1 Flash Image', 'Chọn model gemini-3.1-flash-image-preview, gõ mô tả', 'Ảnh tạo nhanh, chất lượng OK'),
    ('USER', 'Tạo ảnh AI', 'Gemini 3 Pro Image', 'Chọn model gemini-3-pro-image-preview, gõ mô tả', 'Ảnh chất lượng cao, chi tiết tốt'),
    ('USER', 'Tạo ảnh AI', 'Grok Imagine', 'Chọn model grok-imagine-image, gõ mô tả', 'Ảnh tạo đúng mô tả (xAI)'),
    ('USER', 'Tạo ảnh AI', 'Grok Imagine Pro', 'Chọn model grok-imagine-image-pro, gõ mô tả', 'Ảnh chất lượng cao hơn, nhanh hơn'),
    ('USER', 'Tạo ảnh AI', 'Sửa ảnh nhiều lượt', 'L1: Vẽ → L2: Thêm chi tiết → L3: Đổi style', 'Mỗi lượt chỉnh dựa trên ảnh trước'),
    ('USER', 'Tạo ảnh AI', 'Tạo ảnh phức tạp', 'Gõ mô tả dài, nhiều chi tiết cụ thể', 'Ảnh phản ánh đầy đủ mô tả'),
    ('USER', 'Tạo ảnh AI', 'Prompt tiếng Việt', 'Gõ prompt bằng tiếng Việt', 'Hiểu và tạo ảnh đúng'),
    # ──────────────── A3. Gửi file ────────────────
    ('USER', 'Gửi file & truy xuất', 'Upload PDF', 'Gửi file PDF ~5 trang + hỏi nội dung', 'AI đọc và trả lời chính xác'),
    ('USER', 'Gửi file & truy xuất', 'Upload Word', 'Gửi file .docx + hỏi tóm tắt', 'Tóm tắt đúng, đầy đủ ý chính'),
    ('USER', 'Gửi file & truy xuất', 'Upload Excel', 'Gửi file Excel số liệu + hỏi tổng', 'Đọc số liệu, trả lời đúng'),
    ('USER', 'Gửi file & truy xuất', 'Upload CSV', 'Gửi file CSV + hỏi thông tin', 'Đọc đúng, trả lời chính xác'),
    ('USER', 'Gửi file & truy xuất', 'Upload TXT', 'Gửi file .txt + hỏi nội dung', 'Đọc và trả lời đúng'),
    ('USER', 'Gửi file & truy xuất', 'Upload nhiều file', 'Gửi 3 file khác định dạng cùng lúc', 'Xử lý tất cả, trả lời liên quan'),
    ('USER', 'Gửi file & truy xuất', 'Upload file lớn', 'Gửi file PDF ~20-30MB', 'Thành công hoặc báo lỗi rõ ràng'),
    ('USER', 'Gửi file & truy xuất', 'Hỏi đáp sâu file', 'Upload → hỏi 3-4 câu liên tiếp', 'AI nhớ file, trả lời chính xác'),
    ('USER', 'Gửi file & truy xuất', 'Upload HTML', 'Gửi file .html + hỏi', 'Đọc được nội dung text'),
    ('USER', 'Gửi file & truy xuất', 'Trích dẫn nguồn', 'Hỏi từ file đã upload', 'Trích dẫn tên file, trang nguồn'),
    # ──────────────── A4. Chọn Model (4 nhà cung cấp) ────────────────
    ('USER', 'Chọn Model', 'Danh sách model', 'Mở dropdown chọn model', 'Hiển thị đầy đủ tất cả model (4 providers)'),
    # OpenAI
    ('USER', 'Chọn Model', 'GPT-5.4 (OpenAI Flagship)', 'Chọn gpt-5.4 → hỏi 1 câu', 'Trả lời thành công, chất lượng cao nhất'),
    ('USER', 'Chọn Model', 'GPT-5.2 (OpenAI)', 'Chọn gpt-5.2 → hỏi 1 câu', 'Trả lời nhanh, chất lượng tốt'),
    ('USER', 'Chọn Model', 'GPT-5 (OpenAI)', 'Chọn gpt-5 → hỏi 1 câu', 'Trả lời thành công, cân bằng giá/chất lượng'),
    # Google Gemini
    ('USER', 'Chọn Model', 'Gemini 3.1 Pro (Google)', 'Chọn gemini-3.1-pro-preview → hỏi 1 câu', 'Flagship Google, reasoning mạnh, TV tốt'),
    ('USER', 'Chọn Model', 'Gemini 3.1 Flash Lite (Google)', 'Chọn gemini-3.1-flash-lite-preview → hỏi 1 câu', 'Trả lời nhanh, chi phí thấp'),
    ('USER', 'Chọn Model', 'Gemini 2.5 Flash (Google)', 'Chọn gemini-2.5-flash → hỏi 1 câu', 'Trả lời nhanh, chất lượng ổn'),
    # xAI Grok
    ('USER', 'Chọn Model', 'Grok 4.20 Reasoning (xAI)', 'Chọn grok-4.20-reasoning → hỏi 1 câu', 'Reasoning mạnh, trả lời chi tiết'),
    ('USER', 'Chọn Model', 'Grok 4-1 Fast (xAI)', 'Chọn grok-4-1-fast-reasoning → hỏi 1 câu', 'Trả lời nhanh, chi phí thấp'),
    # Anthropic Claude
    ('USER', 'Chọn Model', 'Claude Opus 4.6 (Anthropic)', 'Chọn claude-opus-4-6 → hỏi 1 câu', 'Flagship Anthropic, chất lượng cao'),
    ('USER', 'Chọn Model', 'Claude Sonnet 4.6 (Anthropic)', 'Chọn claude-sonnet-4-6 → hỏi 1 câu', 'Cân bằng giá/chất lượng'),
    ('USER', 'Chọn Model', 'Claude Haiku 4.5 (Anthropic)', 'Chọn claude-haiku-4-5 → hỏi 1 câu', 'Trả lời nhanh, chi phí thấp'),
    # Chức năng model
    ('USER', 'Chọn Model', 'Chuyển model giữa chừng', 'Chat gpt-5.4 → đổi claude-opus-4-6 → hỏi tiếp', 'Hội thoại tiếp tục đúng'),
    ('USER', 'Chọn Model', 'Model mặc định', 'Settings → chọn model mặc định', 'Chat mới dùng model đã chọn'),
    # ──────────────── A5. Knowledge ────────────────
    ('USER', 'Kho kiến thức', 'Tạo Knowledge', 'Workspace → Knowledge → Create', 'Collection tạo thành công'),
    ('USER', 'Kho kiến thức', 'Upload PDF', 'Upload file PDF ~5MB vào Knowledge', 'Upload thành công, được index'),
    ('USER', 'Kho kiến thức', 'Upload Word', 'Upload file .docx vào Knowledge', 'Upload thành công'),
    ('USER', 'Kho kiến thức', 'Upload Excel', 'Upload file .xlsx vào Knowledge', 'Upload thành công'),
    ('USER', 'Kho kiến thức', 'Upload CSV', 'Upload file .csv vào Knowledge', 'Upload thành công'),
    ('USER', 'Kho kiến thức', 'Upload nhiều file', 'Upload 5+ file khác nhau vào 1 Knowledge', 'Tất cả upload thành công'),
    ('USER', 'Kho kiến thức', 'Upload file lớn', 'Upload file ~30-50MB', 'Thành công hoặc báo lỗi rõ'),
    ('USER', 'Kho kiến thức', 'Gọi Knowledge #', 'Trong chat gõ #tên-knowledge', 'Hiển thị gợi ý, chọn được'),
    ('USER', 'Kho kiến thức', 'Hỏi đáp từ Knowledge', 'Gọi # rồi hỏi nội dung', 'Trả lời dựa trên Knowledge, trích dẫn'),
    ('USER', 'Kho kiến thức', 'Xóa file', 'Xóa 1 file trong Knowledge', 'File + embedding bị xóa'),
    ('USER', 'Kho kiến thức', 'Xóa Knowledge', 'Xóa cả collection', 'Collection + tất cả bị xóa'),
    # ──────────────── A6. Workspace ────────────────
    ('USER', 'Không gian làm việc', 'Tạo Workspace', 'Tạo Workspace mới', 'Workspace tạo thành công'),
    ('USER', 'Không gian làm việc', 'Chọn Workspace', 'Chuyển đổi giữa các Workspace', 'Chuyển mượt, dữ liệu riêng'),
    ('USER', 'Không gian làm việc', 'Chia sẻ Knowledge', 'Chia sẻ Knowledge cho user khác', 'User khác truy cập được'),
    ('USER', 'Không gian làm việc', 'Access control', 'Giới hạn quyền truy cập', 'Chỉ user được phép mới truy cập'),
    # ──────────────── A7. Voice ────────────────
    ('USER', 'Giọng nói', 'Text-to-Speech', 'Nhấn icon 🔊 trên response', 'AI đọc bằng giọng nói tự nhiên'),
    ('USER', 'Giọng nói', 'Speech-to-Text', 'Nhấn icon 🎤 → nói câu hỏi', 'Giọng nói → text chính xác'),
    ('USER', 'Giọng nói', 'STT tiếng Việt', 'Nói bằng tiếng Việt', 'Nhận dạng TV chính xác'),
    # ──────────────── A8. Web Search ────────────────
    ('USER', 'Tìm kiếm Web', 'Web search cơ bản', 'Hỏi: "Tin tức mới nhất về AI"', 'Search web, trả lời có nguồn trích dẫn'),
    ('USER', 'Tìm kiếm Web', 'Trích dẫn nguồn web', 'Kiểm tra response có tag nguồn', 'Chip/tag website + URL'),
    ('USER', 'Tìm kiếm Web', 'Search tiếng Việt', 'Hỏi: "Thời tiết Hà Nội hôm nay"', 'Tìm, trả lời TV, có nguồn'),
    # ──────────────── A9. UI ────────────────
    ('USER', 'Giao diện & Cá nhân hóa', 'Dark/Light mode', 'Settings → Theme → chuyển đổi', 'Giao diện chuyển đổi mượt'),
    ('USER', 'Giao diện & Cá nhân hóa', 'Ngôn ngữ', 'Settings → Language → Tiếng Việt', 'UI chuyển tiếng Việt'),
    ('USER', 'Giao diện & Cá nhân hóa', 'System prompt', 'Settings → đặt system prompt', 'AI tuân theo khi chat'),
    ('USER', 'Giao diện & Cá nhân hóa', 'Responsive mobile', 'Mở trên điện thoại', 'Hiển thị tốt, không vỡ layout'),
    # ──────────────── B1. Quản lý user ────────────────
    ('ADMIN', 'Quản lý người dùng', 'Xem danh sách user', 'Admin Panel → Users', 'Hiển thị đầy đủ, role, trạng thái'),
    ('ADMIN', 'Quản lý người dùng', 'Tạo user mới', 'Tạo account / đăng ký', 'Account tạo, trạng thái Pending'),
    ('ADMIN', 'Quản lý người dùng', 'Duyệt user', 'Approve user pending', 'User Active, login được'),
    ('ADMIN', 'Quản lý người dùng', 'Đổi role', 'User → Admin hoặc ngược lại', 'Role + quyền thay đổi'),
    ('ADMIN', 'Quản lý người dùng', 'Xóa user', 'Xóa 1 user account', 'User bị xóa, không login'),
    ('ADMIN', 'Quản lý người dùng', 'Reset password', 'Reset password cho user', 'Password mới hoạt động'),
    # ──────────────── B2. Quản lý model ────────────────
    ('ADMIN', 'Quản lý Model', 'Bật/tắt model', 'Admin → Models → Disable', 'User không thấy model disabled'),
    ('ADMIN', 'Quản lý Model', 'Cấu hình params', 'Set temperature, max_tokens', 'Params áp dụng khi chat'),
    ('ADMIN', 'Quản lý Model', 'Gán Knowledge', 'Gán Knowledge mặc định cho model', 'Model tự động dùng Knowledge'),
    ('ADMIN', 'Quản lý Model', 'Access control', 'Giới hạn model cho user cụ thể', 'Chỉ user được phép mới thấy'),
    # ──────────────── B3. Dashboard ────────────────
    ('ADMIN', 'Dashboard & Giám sát', 'Dashboard tổng quan', 'Truy cập Dashboard middleware', 'Hiển thị tổng chi phí, request, top users'),
    ('ADMIN', 'Dashboard & Giám sát', 'Log requests', 'Xem log chi tiết API request', 'User, model, tokens, cost, timestamp'),
    ('ADMIN', 'Dashboard & Giám sát', 'Báo cáo theo user', 'Lọc chi phí theo user', 'Dữ liệu chính xác'),
    ('ADMIN', 'Dashboard & Giám sát', 'Báo cáo theo model', 'Lọc chi phí theo model', 'Dữ liệu chính xác'),
    ('ADMIN', 'Dashboard & Giám sát', 'Báo cáo theo thời gian', 'Lọc theo khoảng thời gian', 'Dữ liệu đúng khoảng chọn'),
    # ──────────────── B4. Quota ────────────────
    ('ADMIN', 'Quota & Chi phí', 'Set quota user', 'Dashboard → set limit_cost_usd', 'Quota lưu thành công'),
    ('ADMIN', 'Quota & Chi phí', 'Cảnh báo quota', 'User chat gần hết quota', 'Hiển thị cảnh báo'),
    ('ADMIN', 'Quota & Chi phí', 'Chặn khi hết', 'User vượt quota → chat tiếp', 'Từ chối, báo HTTP 429'),
    ('ADMIN', 'Quota & Chi phí', 'Sub-key API', 'Kiểm tra user có sub-key riêng', 'Mỗi user có key riêng'),
    # ──────────────── B5. Cấu hình ────────────────
    ('ADMIN', 'Cấu hình hệ thống', 'WebUI settings', 'Admin → Settings → General', 'Cấu hình lưu và áp dụng'),
    ('ADMIN', 'Cấu hình hệ thống', 'Connections', 'Admin → Settings → Connections', 'Kết nối 4 providers (OpenAI, Gemini, xAI, Anthropic), status OK'),
    ('ADMIN', 'Cấu hình hệ thống', 'RAG config', 'Điều chỉnh chunk size, file limit', 'Thay đổi được áp dụng'),
    ('ADMIN', 'Cấu hình hệ thống', 'Signup control', 'Bật/tắt đăng ký mới', 'Đúng trạng thái bật/tắt'),
]

# Examples for Sheet 3 (updated model names)
BIZ_EXAMPLES = [
    ('Nguyễn Văn A', 'Kế toán', 'Phân tích báo cáo tài chính', 'gpt-5.4', 'Upload file báo cáo Q1.xlsx, hỏi "Tóm tắt doanh thu và chi phí theo tháng"', 'bao_cao_Q1.xlsx', 'AI tổng hợp đúng số liệu, phân tích theo tháng', '', '', 'PASS', '', '', ''),
    ('Trần Thị B', 'Marketing', 'Viết nội dung quảng cáo', 'gemini-3.1-pro-preview', '"Viết 3 mẫu quảng cáo Facebook cho đèn LED thông minh"', '', 'Nội dung sáng tạo, đúng tone, có CTA', '', '', '', '', '', ''),
    ('Lê Văn C', 'R&D', 'Tra cứu tài liệu kỹ thuật', 'claude-opus-4-6', 'Upload datasheet LED → hỏi thông số kỹ thuật', 'LED_Datasheet.pdf', 'Trích xuất đúng thông số từ datasheet', '', '', '', '', '', ''),
    ('Phạm Thị D', 'Nhân sự', 'Soạn thảo văn bản', 'gpt-5', '"Soạn email thông báo chính sách nghỉ phép mới cho toàn công ty"', '', 'Email chuyên nghiệp, đúng format, đầy đủ nội dung', '', '', '', '', '', ''),
    ('Hoàng Văn E', 'IT', 'Debug code', 'grok-4.20-reasoning', 'Gửi đoạn code lỗi Python và hỏi cách sửa', 'error_code.py', 'Phát hiện lỗi, đề xuất fix đúng', '', '', '', '', '', ''),
]


def s(cell, font=None, fill=None, alignment=None, border=None):
    """Apply style to cell."""
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border


def make_title_row(ws, row, text, ncols, font=TITLE_FONT, fill=TITLE_FILL, h=36):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    s(c, font, fill, CENTER)
    for col in range(2, ncols + 1):
        s(ws.cell(row=row, column=col), fill=fill)
    ws.row_dimensions[row].height = h


def make_header_row(ws, row, cols):
    for i, (name, _) in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=name)
        s(c, HEADER_FONT, HEADER_FILL, CENTER, BORDER)
    ws.row_dimensions[row].height = 32


def set_col_widths(ws, cols):
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════
# SHEET 1: Hướng dẫn
# ═══════════════════════════════════════════════════
def build_guide(ws):
    ws.sheet_properties.tabColor = 'FFC000'
    gcols = [('', 3), ('Mục', 20), ('Chi tiết', 60)]
    set_col_widths(ws, gcols)

    row = 1
    make_title_row(ws, row, 'HƯỚNG DẪN SỬ DỤNG FORM KIỂM THỬ', 3)
    row += 2

    # Ký hiệu
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='1. KÝ HIỆU KẾT QUẢ')
    s(c, Font(name='Arial', size=11, bold=True, color='1F4E79'),
      PatternFill('solid', fgColor='D6E4F0'), LEFT_CENTER)
    for col in range(2, 4):
        s(ws.cell(row=row, column=col), fill=PatternFill('solid', fgColor='D6E4F0'))
    row += 1

    symbols = [
        ('✅ PASS', 'Tính năng hoạt động đúng kỳ vọng'),
        ('❌ FAIL', 'Tính năng không hoạt động hoặc sai kỳ vọng'),
        ('⚠️ PARTIAL', 'Hoạt động một phần, cần ghi chú chi tiết'),
        ('⏳ SKIP', 'Chưa test / Bỏ qua'),
    ]
    for sym, desc in symbols:
        ws.cell(row=row, column=2, value=sym).font = BOLD9
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=3, value=desc).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = WRAP
        ws.cell(row=row, column=3).border = BORDER
        row += 1
    row += 1

    # Thang đánh giá
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='2. THANG ĐÁNH GIÁ (1-5)')
    s(c, Font(name='Arial', size=11, bold=True, color='1F4E79'),
      PatternFill('solid', fgColor='D6E4F0'), LEFT_CENTER)
    for col in range(2, 4):
        s(ws.cell(row=row, column=col), fill=PatternFill('solid', fgColor='D6E4F0'))
    row += 1

    ratings = [
        ('1 - Rất kém', 'Không hoạt động, lỗi nghiêm trọng'),
        ('2 - Kém', 'Hoạt động nhưng nhiều lỗi, khó dùng'),
        ('3 - Trung bình', 'Hoạt động cơ bản, cần cải thiện'),
        ('4 - Tốt', 'Hoạt động tốt, ít lỗi nhỏ'),
        ('5 - Xuất sắc', 'Hoạt động hoàn hảo, trải nghiệm tuyệt vời'),
    ]
    for r, desc in ratings:
        ws.cell(row=row, column=2, value=r).font = BOLD9
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=3, value=desc).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = WRAP
        ws.cell(row=row, column=3).border = BORDER
        row += 1
    row += 1

    # Cột cần điền
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='3. CÁC CỘT NGƯỜI TEST CẦN ĐIỀN')
    s(c, Font(name='Arial', size=11, bold=True, color='1F4E79'),
      PatternFill('solid', fgColor='D6E4F0'), LEFT_CENTER)
    for col in range(2, 4):
        s(ws.cell(row=row, column=col), fill=PatternFill('solid', fgColor='D6E4F0'))
    row += 1

    fill_cols = [
        ('Kết quả thực tế', 'Ghi lại response AI hoặc kết quả thao tác thực tế'),
        ('Ảnh chụp màn hình', 'Ghi tên file ảnh (VD: TC01_chat.png) hoặc paste link ảnh'),
        ('Kết quả', 'Chọn từ dropdown: PASS / FAIL / PARTIAL / SKIP'),
        ('Đánh giá (1-5)', 'Chọn từ dropdown: 1 đến 5'),
        ('Thời gian test', 'Điền ngày test theo format dd/mm/yyyy'),
        ('Nhận xét, đánh giá', 'Mô tả chi tiết bug, issue, hoặc nhận xét'),
        ('Feedback / Đề xuất', 'Ý kiến cải thiện, đề xuất nâng cấp tính năng'),
    ]
    for k, v in fill_cols:
        ws.cell(row=row, column=2, value=k).font = BOLD9
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=3, value=v).font = DATA_FONT
        ws.cell(row=row, column=3).alignment = WRAP
        ws.cell(row=row, column=3).border = BORDER
        row += 1
    row += 1

    # Quy trình test
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='4. QUY TRÌNH TEST')
    s(c, Font(name='Arial', size=11, bold=True, color='1F4E79'),
      PatternFill('solid', fgColor='D6E4F0'), LEFT_CENTER)
    for col in range(2, 4):
        s(ws.cell(row=row, column=col), fill=PatternFill('solid', fgColor='D6E4F0'))
    row += 1

    steps = [
        'Bước 1: Đọc hướng dẫn (sheet này)',
        'Bước 2: Truy cập https://openwebui.rangdong.com.vn',
        'Bước 3: Thực hiện từng test case theo sheet "Form Test Tính Năng"',
        'Bước 4: Điền kết quả, chụp ảnh màn hình nếu có',
        'Bước 5: Ghi nhận xét và feedback cho từng mục',
        'Bước 6: Test thêm theo nghiệp vụ riêng ở sheet "Form Test Nghiệp Vụ"',
        'Bước 7: Gửi file Excel đã điền về cho team phát triển',
    ]
    for step in steps:
        ws.cell(row=row, column=2, value=step).font = DATA_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=2).alignment = LEFT_CENTER
        row += 1


# ═══════════════════════════════════════════════════
# SHEET 2: Form Test Tính Năng
# ═══════════════════════════════════════════════════
def build_test_form(ws):
    ws.sheet_properties.tabColor = '1F4E79'
    set_col_widths(ws, COLS2)
    ncols = len(COLS2)

    row = 1
    make_title_row(ws, row, 'FORM KIỂM THỬ TÍNH NĂNG - OPEN WEBUI', ncols)
    row += 1
    # Info row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1,
            value=f'Phiên bản: v2.1  |  Ngày: 06/04/2026  |  URL: https://openwebui.rangdong.com.vn  |  Tổng: {len(TEST_CASES)} test cases  |  4 providers: OpenAI, Google, xAI, Anthropic')
    s(ws.cell(row=row, column=1), Font(name='Arial', size=9, italic=True, color='666666'), alignment=CENTER)
    row += 1

    # Header
    header_row = row
    make_header_row(ws, row, COLS2)
    row += 1
    data_start = row

    # Track groups for summary
    groups_order = []
    seen_groups = set()

    # Insert data
    current_phanhe = None
    group_idx = 0
    current_group = None

    for stt_i, tc in enumerate(TEST_CASES, 1):
        phanhe, nhom, tinh_nang, thao_tac, ky_vong = tc

        # Section separator for USER/ADMIN
        if phanhe != current_phanhe:
            current_phanhe = phanhe
            label = 'PHẦN A: TÍNH NĂNG NGƯỜI DÙNG (USER)' if phanhe == 'USER' else 'PHẦN B: TÍNH NĂNG QUẢN TRỊ (ADMIN)'
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            c = ws.cell(row=row, column=1, value=label)
            s(c, SECTION_FONT, SECTION_FILLS[phanhe], LEFT_CENTER)
            for col in range(2, ncols + 1):
                s(ws.cell(row=row, column=col), fill=SECTION_FILLS[phanhe])
            ws.row_dimensions[row].height = 28
            row += 1

        # Track group color
        if nhom != current_group:
            current_group = nhom
            group_idx += 1
        if nhom not in seen_groups:
            seen_groups.add(nhom)
            groups_order.append((phanhe, nhom))

        fill = PatternFill('solid', fgColor=GROUP_COLORS[(group_idx - 1) % len(GROUP_COLORS)])

        vals = [stt_i, phanhe, nhom, tinh_nang, thao_tac, ky_vong, '', '', '', '', '', '', '']
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=v)
            al = CENTER if i in (1, 2, 9, 10) else WRAP
            s(c, DATA_FONT, fill, al, BORDER)

        ws.row_dimensions[row].height = 40
        row += 1

    data_end = row - 1

    # Freeze pane below header
    ws.freeze_panes = f'A{header_row + 1}'

    # Auto-filter
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(ncols)}{data_end}'

    # Data validation: Kết quả (col 9)
    dv_result = DataValidation(type='list', formula1='"PASS,FAIL,PARTIAL,SKIP"', allow_blank=True)
    dv_result.error = 'Chọn PASS, FAIL, PARTIAL hoặc SKIP'
    dv_result.errorTitle = 'Giá trị không hợp lệ'
    dv_result.prompt = 'Chọn kết quả'
    dv_result.promptTitle = 'Kết quả test'
    ws.add_data_validation(dv_result)
    dv_result.add(f'I{header_row + 1}:I{data_end + 50}')

    # Data validation: Đánh giá (col 10)
    dv_rating = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
    dv_rating.prompt = 'Chọn 1-5'
    ws.add_data_validation(dv_rating)
    dv_rating.add(f'J{header_row + 1}:J{data_end + 50}')

    # Conditional formatting on column I (Kết quả)
    rng = f'I{header_row + 1}:I{data_end + 50}'
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"PASS"'], fill=PASS_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"FAIL"'], fill=FAIL_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"PARTIAL"'], fill=PARTIAL_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"SKIP"'], fill=SKIP_FILL))

    # Date format for col K (Thời gian test)
    for r in range(header_row + 1, data_end + 51):
        ws.cell(row=r, column=11).number_format = 'DD/MM/YYYY'

    return groups_order, header_row, data_end


# ═══════════════════════════════════════════════════
# SHEET 3: Form Test Nghiệp Vụ
# ═══════════════════════════════════════════════════
def build_biz_form(ws):
    ws.sheet_properties.tabColor = '548235'
    set_col_widths(ws, COLS3)
    ncols = len(COLS3)

    row = 1
    make_title_row(ws, row, 'FORM KIỂM THỬ THEO NGHIỆP VỤ', ncols)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1,
            value='Mỗi anh/chị tự điền kịch bản test theo nghiệp vụ riêng. Dưới đây là 5 ví dụ mẫu.')
    s(ws.cell(row=row, column=1), Font(name='Arial', size=9, italic=True, color='666666'), alignment=CENTER)
    row += 1

    # Header
    header_row = row
    make_header_row(ws, row, COLS3)
    row += 1

    # Example rows
    for idx, ex in enumerate(BIZ_EXAMPLES, 1):
        vals = [idx] + list(ex)
        fill = PatternFill('solid', fgColor='FFF8E7') if idx % 2 == 1 else PatternFill('solid', fgColor='FFFFFF')
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=v)
            al = CENTER if i in (1, 11) else WRAP
            s(c, Font(name='Arial', size=9, italic=True, color='808080'), fill, al, BORDER)
        ws.row_dimensions[row].height = 45
        row += 1

    # Separator
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value='↓↓↓  ĐIỀN TEST CASE CỦA BẠN TỪ ĐÂY  ↓↓↓')
    s(c, Font(name='Arial', size=10, bold=True, color='548235'),
      PatternFill('solid', fgColor='E2EFDA'), CENTER)
    for col in range(2, ncols + 1):
        s(ws.cell(row=row, column=col), fill=PatternFill('solid', fgColor='E2EFDA'))
    ws.row_dimensions[row].height = 28
    row += 1

    # 50 blank rows with format
    blank_start = row
    for idx in range(50):
        stt = len(BIZ_EXAMPLES) + idx + 1
        fill = PatternFill('solid', fgColor='F9F9F9') if idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        ws.cell(row=row, column=1, value=stt)
        for col in range(1, ncols + 1):
            c = ws.cell(row=row, column=col)
            al = CENTER if col in (1, 11) else WRAP
            s(c, DATA_FONT, fill, al, BORDER)
        ws.cell(row=row, column=12).number_format = 'DD/MM/YYYY'
        ws.row_dimensions[row].height = 40
        row += 1
    blank_end = row - 1

    # Freeze + filter
    ws.freeze_panes = f'A{header_row + 1}'
    ws.auto_filter.ref = f'A{header_row}:{get_column_letter(ncols)}{blank_end}'

    # Validation
    dv = DataValidation(type='list', formula1='"PASS,FAIL,PARTIAL,SKIP"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'K{header_row + 1}:K{blank_end}')

    # Conditional formatting
    rng = f'K{header_row + 1}:K{blank_end}'
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"PASS"'], fill=PASS_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"FAIL"'], fill=FAIL_FILL))
    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"PARTIAL"'], fill=PARTIAL_FILL))


# ═══════════════════════════════════════════════════
# SHEET 4: Tổng hợp
# ═══════════════════════════════════════════════════
def build_summary(ws, groups_order, header_row, data_end):
    ws.sheet_properties.tabColor = 'BF4B28'
    scols = [('STT', 5), ('Phân hệ', 10), ('Nhóm tính năng', 28), ('Tổng TC', 10),
             ('PASS', 10), ('FAIL', 10), ('PARTIAL', 10), ('SKIP', 10), ('Tỷ lệ PASS (%)', 14)]
    set_col_widths(ws, scols)
    ncols = len(scols)

    row = 1
    make_title_row(ws, row, 'BẢNG TỔNG HỢP KẾT QUẢ KIỂM THỬ', ncols)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value='Dữ liệu tự động tính từ sheet "Form Test Tính Năng" bằng COUNTIFS')
    s(ws.cell(row=row, column=1), Font(name='Arial', size=9, italic=True, color='666666'), alignment=CENTER)
    row += 1

    # Header
    for i, (name, _) in enumerate(scols, 1):
        c = ws.cell(row=row, column=i, value=name)
        s(c, HEADER_FONT, HEADER_FILL, CENTER, BORDER)
    ws.row_dimensions[row].height = 28
    s_header = row
    row += 1

    # Sheet ref
    sn = "'Form Test Tính Năng'"
    col_nhom = 'C'   # nhóm tính năng column
    col_kq = 'I'     # kết quả column
    # Data range in the test form sheet (skip section separator rows)
    dr_start = header_row + 1
    dr_end = data_end + 50

    for idx, (phanhe, nhom) in enumerate(groups_order, 1):
        fill = PatternFill('solid', fgColor='F2F7FB') if idx % 2 == 1 else PatternFill('solid', fgColor='FFFFFF')

        ws.cell(row=row, column=1, value=idx)
        s(ws.cell(row=row, column=1), DATA_FONT, fill, CENTER, BORDER)

        ws.cell(row=row, column=2, value=phanhe)
        s(ws.cell(row=row, column=2), DATA_FONT, fill, CENTER, BORDER)

        ws.cell(row=row, column=3, value=nhom)
        s(ws.cell(row=row, column=3), DATA_FONT, fill, WRAP, BORDER)

        # COUNTIFS formulas
        nhom_esc = nhom.replace('"', '""')
        base = f'COUNTIFS({sn}!{col_nhom}${dr_start}:{col_nhom}${dr_end},"{nhom_esc}"'

        # Tổng TC (count all rows with this group name)
        ws.cell(row=row, column=4).value = f'={base})'
        s(ws.cell(row=row, column=4), BOLD9, fill, CENTER, BORDER)

        # PASS
        ws.cell(row=row, column=5).value = f'={base},{sn}!{col_kq}${dr_start}:{col_kq}${dr_end},"PASS")'
        s(ws.cell(row=row, column=5), DATA_FONT, PASS_FILL, CENTER, BORDER)

        # FAIL
        ws.cell(row=row, column=6).value = f'={base},{sn}!{col_kq}${dr_start}:{col_kq}${dr_end},"FAIL")'
        s(ws.cell(row=row, column=6), DATA_FONT, FAIL_FILL, CENTER, BORDER)

        # PARTIAL
        ws.cell(row=row, column=7).value = f'={base},{sn}!{col_kq}${dr_start}:{col_kq}${dr_end},"PARTIAL")'
        s(ws.cell(row=row, column=7), DATA_FONT, PARTIAL_FILL, CENTER, BORDER)

        # SKIP
        ws.cell(row=row, column=8).value = f'={base},{sn}!{col_kq}${dr_start}:{col_kq}${dr_end},"SKIP")'
        s(ws.cell(row=row, column=8), DATA_FONT, SKIP_FILL, CENTER, BORDER)

        # Tỷ lệ PASS
        d4 = get_column_letter(4)
        e = get_column_letter(5)
        ws.cell(row=row, column=9).value = f'=IF({d4}{row}=0,"",{e}{row}/{d4}{row})'
        ws.cell(row=row, column=9).number_format = '0.0%'
        s(ws.cell(row=row, column=9), BOLD9, fill, CENTER, BORDER)

        ws.row_dimensions[row].height = 24
        row += 1

    # Total row
    total_row = row
    ws.cell(row=row, column=1, value='')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value='TỔNG CỘNG')
    total_fill = PatternFill('solid', fgColor='D6E4F0')
    tfont = Font(name='Arial', size=10, bold=True, color='1F4E79')
    s(ws.cell(row=row, column=1), tfont, total_fill, CENTER, BORDER)
    for col in range(2, 4):
        s(ws.cell(row=row, column=col), fill=total_fill, border=BORDER)

    first_data = s_header + 1
    for col in range(4, 9):
        letter = get_column_letter(col)
        ws.cell(row=row, column=col).value = f'=SUM({letter}{first_data}:{letter}{total_row - 1})'
        s(ws.cell(row=row, column=col), tfont, total_fill, CENTER, BORDER)

    # Total pass rate
    ws.cell(row=row, column=9).value = f'=IF(D{row}=0,"",E{row}/D{row})'
    ws.cell(row=row, column=9).number_format = '0.0%'
    s(ws.cell(row=row, column=9), tfont, total_fill, CENTER, BORDER)
    ws.row_dimensions[row].height = 30
    row += 2

    # ── Thông tin người test ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value='THÔNG TIN NGƯỜI TEST')
    s(c, Font(name='Arial', size=11, bold=True, color='1F4E79'),
      PatternFill('solid', fgColor='D6E4F0'), LEFT_CENTER)
    for col in range(2, ncols + 1):
        s(ws.cell(row=row, column=col), fill=PatternFill('solid', fgColor='D6E4F0'))
    row += 1

    info_fields = ['Họ và tên', 'Phòng ban', 'Ngày test', 'Trình duyệt (Chrome/Edge/Firefox)',
                   'Thiết bị (PC/Laptop/Mobile)', 'Nhận xét chung', 'Đề xuất cải thiện tổng thể']
    for field in info_fields:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=field)
        s(ws.cell(row=row, column=1), BOLD9, border=BORDER, alignment=LEFT_CENTER)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=ncols)
        ws.cell(row=row, column=4, value='')
        s(ws.cell(row=row, column=4), border=BORDER, alignment=WRAP)
        for col in range(5, ncols + 1):
            s(ws.cell(row=row, column=col), border=BORDER)
        ws.row_dimensions[row].height = 30
        row += 1


# ═══════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'Hướng dẫn'
    build_guide(ws1)

    ws2 = wb.create_sheet('Form Test Tính Năng')
    groups_order, header_row, data_end = build_test_form(ws2)

    ws3 = wb.create_sheet('Form Test Nghiệp Vụ')
    build_biz_form(ws3)

    ws4 = wb.create_sheet('Tổng hợp')
    build_summary(ws4, groups_order, header_row, data_end)

    out = 'FORM_TEST_OPEN_WEBUI_v2.xlsx'
    wb.save(out)
    print(f'✅ Đã tạo: {out}')
    print(f'   📋 Sheet 1: Hướng dẫn')
    print(f'   🧪 Sheet 2: Form Test Tính Năng - 1 bảng, {len(TEST_CASES)} test cases, 13 cột')
    print(f'   📝 Sheet 3: Form Test Nghiệp Vụ - {len(BIZ_EXAMPLES)} ví dụ + 50 dòng trống')
    print(f'   📊 Sheet 4: Tổng hợp - COUNTIFS tự động + thông tin người test')


if __name__ == '__main__':
    main()
